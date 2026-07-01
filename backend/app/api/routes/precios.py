"""
precios.py — catálogo de precios de supermercados uruguayos.

Endpoints públicos (requieren JWT de usuario):
  GET  /precios              — lista paginada con filtros
  GET  /precios/tiendas      — tiendas disponibles
  GET  /precios/categorias   — categorías disponibles
  GET  /precios/{id}         — detalle de un producto

Endpoint de sincronización (requiere APP_SECRET_KEY en header X-Sync-Key):
  POST /precios/sync         — bulk upsert desde el scraper (INSERT ... ON CONFLICT)
"""

import csv
import io
import logging
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from fastapi.security import APIKeyHeader
from sqlalchemy import select, func
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.database import get_db
from app.core.deps import get_current_user, require_permission
from app.models.producto import Producto
from app.models.user import User
from app.schemas.precios import (
    ProductoOut, SyncPayload, SyncResult, PreciosListResponse,
    CompararResponse, CompararGrupo, CompararTiendaItem,
    PreciosStats, TiendaStats,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/precios", tags=["precios"])

_sync_key_scheme = APIKeyHeader(name="X-Sync-Key", auto_error=False)


async def _require_sync_key(key: Optional[str] = Depends(_sync_key_scheme)):
    if not key or key != settings.APP_SECRET_KEY:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Sync key inválida")


# ---------------------------------------------------------------------------
# Sync endpoint — llamado por el scraper
# ---------------------------------------------------------------------------

@router.post("/sync", response_model=SyncResult, dependencies=[Depends(_require_sync_key)])
async def sync_productos(payload: SyncPayload, db: AsyncSession = Depends(get_db)):
    """Bulk upsert via PostgreSQL INSERT ... ON CONFLICT DO UPDATE.
    Un solo statement por batch — eficiente para 45k+ productos.
    """
    if not payload.productos:
        return SyncResult(upsertados=0, total_enviados=0)

    now = datetime.now(timezone.utc)
    rows = []
    for item in payload.productos:
        ts = None
        if item.actualizado_en:
            try:
                ts = datetime.fromisoformat(item.actualizado_en)
            except ValueError:
                ts = now
        rows.append({
            "tienda":        item.tienda,
            "url":           item.url,
            "nombre":        item.nombre,
            "precio":        item.precio,
            "precio_lista":  item.precio_lista,
            "sku":           item.sku,
            "barcode":       item.barcode,
            "marca":         item.marca,
            "categoria":     item.categoria,
            "actualizado_en": ts or now,
        })

    stmt = pg_insert(Producto).values(rows)
    stmt = stmt.on_conflict_do_update(
        index_elements=["url"],
        set_={
            "tienda":        stmt.excluded.tienda,
            "nombre":        stmt.excluded.nombre,
            "precio":        stmt.excluded.precio,
            "precio_lista":  stmt.excluded.precio_lista,
            "sku":           stmt.excluded.sku,
            "barcode":       stmt.excluded.barcode,
            "marca":         stmt.excluded.marca,
            "categoria":     stmt.excluded.categoria,
            "actualizado_en": stmt.excluded.actualizado_en,
        },
    )

    await db.execute(stmt)
    logger.info("sync_productos: %d upsertados", len(rows))
    return SyncResult(upsertados=len(rows), total_enviados=len(payload.productos))


# ---------------------------------------------------------------------------
# Endpoints de lectura
# ---------------------------------------------------------------------------

@router.get("/tiendas", response_model=list[str])
async def listar_tiendas(
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    rows = await db.execute(
        select(Producto.tienda).distinct().order_by(Producto.tienda)
    )
    return [r[0] for r in rows.all()]


@router.get("/estadisticas", response_model=PreciosStats)
async def estadisticas_precios(
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Resumen estadístico del catálogo de precios por tienda."""
    from sqlalchemy import case, Integer, cast
    desc_expr = cast(
        case(
            (
                Producto.precio_lista.isnot(None) & (Producto.precio_lista > Producto.precio),
                1,
            ),
            else_=0,
        ),
        Integer,
    )
    rows = await db.execute(
        select(
            Producto.tienda,
            func.count().label("total"),
            func.sum(desc_expr).label("con_descuento"),
            func.avg(Producto.precio).label("precio_promedio"),
        ).group_by(Producto.tienda).order_by(Producto.tienda)
    )
    tiendas = []
    total_global = 0
    for tienda, total, con_desc, avg_precio in rows.all():
        tiendas.append(TiendaStats(
            tienda=tienda,
            total=total,
            con_descuento=int(con_desc or 0),
            precio_promedio=round(float(avg_precio), 2) if avg_precio else None,
        ))
        total_global += total
    return PreciosStats(total_productos=total_global, tiendas=tiendas)


@router.get("/categorias", response_model=list[str])
async def listar_categorias(
    tienda: Optional[str] = None,
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    q = select(Producto.categoria).distinct().where(Producto.categoria.isnot(None))
    if tienda:
        q = q.where(Producto.tienda == tienda)
    rows = await db.execute(q.order_by(Producto.categoria))
    return [r[0] for r in rows.all()]


@router.get("", response_model=PreciosListResponse)
async def listar_precios(
    tienda:     Optional[str]   = Query(None),
    categoria:  Optional[str]   = Query(None),
    marca:      Optional[str]   = Query(None),
    q:          Optional[str]   = Query(None, description="Búsqueda por nombre, SKU o barcode"),
    precio_min:    Optional[float] = Query(None),
    precio_max:    Optional[float] = Query(None),
    con_descuento: Optional[bool]  = Query(None, description="Solo productos con descuento activo"),
    sort_by:       Optional[str]   = Query(None, description="Campo de ordenamiento: nombre|precio|tienda|categoria"),
    sort_dir:      Optional[str]   = Query("asc", description="Dirección: asc|desc"),
    page:          int             = Query(1, ge=1),
    page_size:     int             = Query(50, ge=1, le=200),
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    base    = select(Producto)
    count_q = select(func.count()).select_from(Producto)

    filters = []
    if tienda:
        filters.append(Producto.tienda == tienda)
    if categoria:
        filters.append(Producto.categoria.ilike(f"%{categoria}%"))
    if marca:
        filters.append(Producto.marca.ilike(f"%{marca}%"))
    if q:
        like = f"%{q}%"
        filters.append(
            Producto.nombre.ilike(like) |
            Producto.sku.ilike(like)    |
            Producto.barcode.ilike(like)
        )
    if precio_min is not None:
        filters.append(Producto.precio >= precio_min)
    if precio_max is not None:
        filters.append(Producto.precio <= precio_max)
    if con_descuento:
        filters.append(
            Producto.precio_lista.isnot(None) &
            (Producto.precio_lista > Producto.precio)
        )

    for f in filters:
        base    = base.where(f)
        count_q = count_q.where(f)

    total_result = await db.execute(count_q)
    total = total_result.scalar_one()

    _sort_cols = {
        "nombre":    Producto.nombre,
        "precio":    Producto.precio,
        "tienda":    Producto.tienda,
        "categoria": Producto.categoria,
    }
    sort_col = _sort_cols.get(sort_by or "nombre", Producto.nombre)
    order_expr = sort_col.desc() if sort_dir == "desc" else sort_col.asc()

    offset = (page - 1) * page_size
    items_result = await db.execute(
        base.order_by(order_expr).offset(offset).limit(page_size)
    )
    items = items_result.scalars().all()

    return PreciosListResponse(
        total=total,
        page=page,
        page_size=page_size,
        items=[ProductoOut.model_validate(p) for p in items],
    )


@router.get("/comparar", response_model=CompararResponse)
async def comparar_precios(
    q:       Optional[str] = Query(None, description="Búsqueda por nombre o barcode"),
    barcode: Optional[str] = Query(None, description="Barcode exacto"),
    limit:   int           = Query(30, ge=1, le=100),
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Compara el mismo producto (por barcode) en distintas tiendas.
    Agrupa productos con barcode compartido entre >=2 tiendas.
    """
    base = select(Producto).where(Producto.barcode.isnot(None), Producto.barcode != "")

    if barcode:
        base = base.where(Producto.barcode == barcode)
    elif q:
        like = f"%{q}%"
        base = base.where(
            Producto.nombre.ilike(like) | Producto.barcode.ilike(like)
        )
    else:
        return CompararResponse(grupos=[], total=0)

    result = await db.execute(base.order_by(Producto.barcode, Producto.tienda))
    rows = result.scalars().all()

    from collections import defaultdict
    by_barcode: dict[str, list] = defaultdict(list)
    for p in rows:
        by_barcode[p.barcode].append(p)

    grupos = []
    for bc, prods in by_barcode.items():
        tiendas_set = {p.tienda for p in prods}
        if len(tiendas_set) < 2:
            continue
        nombre_ref = next((p.nombre for p in prods if p.nombre), None)
        items = [
            CompararTiendaItem(
                tienda=p.tienda,
                precio=p.precio,
                precio_lista=p.precio_lista,
                url=p.url,
                nombre=p.nombre,
            )
            for p in sorted(prods, key=lambda x: x.tienda)
        ]
        grupos.append(CompararGrupo(
            barcode=bc,
            nombre_ref=nombre_ref,
            n_tiendas=len(tiendas_set),
            tiendas=items,
        ))

    grupos.sort(key=lambda g: -g.n_tiendas)
    grupos = grupos[:limit]
    return CompararResponse(grupos=grupos, total=len(grupos))


@router.get("/export.csv")
async def exportar_csv(
    tienda:     Optional[str]   = Query(None),
    categoria:  Optional[str]   = Query(None),
    marca:      Optional[str]   = Query(None),
    q:          Optional[str]   = Query(None),
    precio_min: Optional[float] = Query(None),
    precio_max: Optional[float] = Query(None),
    con_descuento: Optional[bool] = Query(None),
    limit:      int             = Query(10000, ge=1, le=50000),
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Exporta el catálogo filtrado como CSV."""
    base = select(Producto)
    filters = []
    if tienda:        filters.append(Producto.tienda == tienda)
    if categoria:     filters.append(Producto.categoria.ilike(f"%{categoria}%"))
    if marca:         filters.append(Producto.marca.ilike(f"%{marca}%"))
    if q:
        like = f"%{q}%"
        filters.append(
            Producto.nombre.ilike(like) | Producto.sku.ilike(like) | Producto.barcode.ilike(like)
        )
    if precio_min is not None: filters.append(Producto.precio >= precio_min)
    if precio_max is not None: filters.append(Producto.precio <= precio_max)
    if con_descuento:
        filters.append(Producto.precio_lista.isnot(None) & (Producto.precio_lista > Producto.precio))
    for f in filters:
        base = base.where(f)

    result = await db.execute(base.order_by(Producto.nombre).limit(limit))
    items = result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["tienda","nombre","precio","precio_lista","sku","barcode","marca","categoria","url"])
    for p in items:
        writer.writerow([
            p.tienda, p.nombre, p.precio, p.precio_lista,
            p.sku, p.barcode, p.marca, p.categoria, p.url,
        ])

    filename = f"precios_{datetime.now(timezone.utc).strftime('%Y-%m-%d')}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/historial/exportar-excel")
async def exportar_excel_historial(
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Exporta historial completo como Excel — una hoja por fecha de scan.

    Usa write_only=True + stream_results para memoria constante (~20MB) sin importar
    cuántas filas haya. Sin esto openpyxl acumula ~1M cell objects en RAM y crashea.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.cell import WriteOnlyCell
    from app.models.precio_historial import PrecioHistorial

    fechas_result = await db.execute(
        select(PrecioHistorial.fecha_scan).distinct().order_by(PrecioHistorial.fecha_scan.desc())
    )
    fechas = [r[0] for r in fechas_result.all()]

    if not fechas:
        raise HTTPException(status_code=404, detail="No hay datos históricos disponibles")

    COL_HEADERS = ["Tienda", "Nombre", "Precio", "Precio Lista", "SKU", "Barcode", "Marca", "Categoría", "URL"]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    center      = Alignment(horizontal="center")

    wb = Workbook(write_only=True)

    for fecha in fechas:
        ws = wb.create_sheet(title=str(fecha))

        # Fila de cabecera con estilo usando WriteOnlyCell
        header_row = []
        for h in COL_HEADERS:
            c = WriteOnlyCell(ws, value=h)
            c.font  = header_font
            c.fill  = header_fill
            c.alignment = center
            header_row.append(c)
        ws.append(header_row)

        result = await db.stream(
            select(
                PrecioHistorial.tienda, PrecioHistorial.nombre,
                PrecioHistorial.precio, PrecioHistorial.precio_lista,
                PrecioHistorial.sku, PrecioHistorial.barcode,
                PrecioHistorial.marca, PrecioHistorial.categoria,
                PrecioHistorial.url,
            )
            .where(PrecioHistorial.fecha_scan == fecha)
            .order_by(PrecioHistorial.tienda, PrecioHistorial.nombre)
            .execution_options(yield_per=500)
        )
        async for row in result:
            ws.append(list(row))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"precios_historial_{datetime.now(timezone.utc).strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/historial/fechas", response_model=list[str])
async def historial_fechas(
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Fechas con snapshots disponibles, ordenadas más reciente primero."""
    from app.models.precio_historial import PrecioHistorial
    rows = await db.execute(
        select(PrecioHistorial.fecha_scan).distinct().order_by(PrecioHistorial.fecha_scan.desc())
    )
    return [str(r[0]) for r in rows.all()]


@router.get("/historial")
async def historial_precios(
    fecha:         str             = Query(..., description="YYYY-MM-DD"),
    tienda:        Optional[str]   = Query(None),
    categoria:     Optional[str]   = Query(None),
    marca:         Optional[str]   = Query(None),
    q:             Optional[str]   = Query(None),
    precio_min:    Optional[float] = Query(None),
    precio_max:    Optional[float] = Query(None),
    con_descuento: Optional[bool]  = Query(None),
    sort_by:       Optional[str]   = Query(None),
    sort_dir:      Optional[str]   = Query("asc"),
    page:          int             = Query(1, ge=1),
    page_size:     int             = Query(50, ge=1, le=200),
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Catálogo de precios de una fecha específica (snapshot histórico)."""
    from datetime import date as date_type
    from app.models.precio_historial import PrecioHistorial

    try:
        fecha_dt = date_type.fromisoformat(fecha)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido (usar YYYY-MM-DD)")

    base    = select(PrecioHistorial).where(PrecioHistorial.fecha_scan == fecha_dt)
    count_q = select(func.count()).select_from(PrecioHistorial).where(PrecioHistorial.fecha_scan == fecha_dt)

    filters = []
    if tienda:     filters.append(PrecioHistorial.tienda == tienda)
    if categoria:  filters.append(PrecioHistorial.categoria.ilike(f"%{categoria}%"))
    if marca:      filters.append(PrecioHistorial.marca.ilike(f"%{marca}%"))
    if q:
        like = f"%{q}%"
        filters.append(
            PrecioHistorial.nombre.ilike(like) |
            PrecioHistorial.sku.ilike(like)    |
            PrecioHistorial.barcode.ilike(like)
        )
    if precio_min is not None: filters.append(PrecioHistorial.precio >= precio_min)
    if precio_max is not None: filters.append(PrecioHistorial.precio <= precio_max)
    if con_descuento:
        filters.append(
            PrecioHistorial.precio_lista.isnot(None) &
            (PrecioHistorial.precio_lista > PrecioHistorial.precio)
        )

    for f in filters:
        base    = base.where(f)
        count_q = count_q.where(f)

    total = (await db.execute(count_q)).scalar_one()

    _sort_cols = {
        "nombre":    PrecioHistorial.nombre,
        "precio":    PrecioHistorial.precio,
        "tienda":    PrecioHistorial.tienda,
        "categoria": PrecioHistorial.categoria,
    }
    sort_col   = _sort_cols.get(sort_by or "nombre", PrecioHistorial.nombre)
    order_expr = sort_col.desc() if sort_dir == "desc" else sort_col.asc()

    offset = (page - 1) * page_size
    items  = (await db.execute(base.order_by(order_expr).offset(offset).limit(page_size))).scalars().all()

    return {
        "fecha":     str(fecha_dt),
        "total":     total,
        "page":      page,
        "page_size": page_size,
        "items": [
            {
                "id":           p.id,
                "tienda":       p.tienda,
                "url":          p.url,
                "nombre":       p.nombre,
                "precio":       p.precio,
                "precio_lista": p.precio_lista,
                "sku":          p.sku,
                "barcode":      p.barcode,
                "marca":        p.marca,
                "categoria":    p.categoria,
                "fecha_scan":   str(p.fecha_scan),
            }
            for p in items
        ],
    }


@router.get("/scraper/status")
async def scraper_status(_: User = Depends(get_current_user)):
    """Estado del scheduler nocturno de scraping."""
    from app.services.scraper_sync import get_status
    return await get_status()


@router.get("/scraper/progress")
async def scraper_progress(_: User = Depends(get_current_user)):
    """Progreso en tiempo real del scan activo (lee JSON de checkpoint)."""
    from app.services.scraper_sync import get_progress
    return await get_progress()


@router.post("/scraper/trigger", status_code=202)
async def scraper_trigger(_: User = Depends(get_current_user)):
    """Dispara un scraping manual completo (todas las tiendas)."""
    from app.services.scraper_sync import trigger_manual
    launched = await trigger_manual()
    if not launched:
        raise HTTPException(status_code=409, detail="Ya hay un scraping en curso")
    return {"message": "Scraping completo iniciado"}


@router.post("/scraper/trigger-gdu", status_code=202)
async def scraper_trigger_gdu(_: User = Depends(get_current_user)):
    """Dispara un scan de solo GDU (Geant, Disco, Devoto). Actualiza únicamente
    registros GDU en PostgreSQL via upsert por URL — no toca Tata ni Farmashop."""
    from app.services.scraper_sync import trigger_gdu
    launched = await trigger_gdu()
    if not launched:
        raise HTTPException(status_code=409, detail="Ya hay un scraping en curso")
    return {"message": "GDU scan iniciado"}


@router.post("/scraper/trigger-botiga", status_code=202)
async def scraper_trigger_botiga(_: User = Depends(get_current_user)):
    """Dispara un scan de solo Botiga (botiga.farmashop.com.uy, Magento GraphQL).
    Upsert por URL en PostgreSQL — no toca otras tiendas."""
    from app.services.scraper_sync import trigger_botiga
    launched = await trigger_botiga()
    if not launched:
        raise HTTPException(status_code=409, detail="Ya hay un scraping en curso")
    return {"message": "Botiga scan iniciado"}


@router.delete("/vaciar", status_code=200)
async def vaciar_catalogo(
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Borra todos los registros de productos y su historial. Solo superusuarios."""
    from sqlalchemy import text
    from app.models.precio_historial import PrecioHistorial
    if not _.is_superuser:
        raise HTTPException(status_code=403, detail="Solo superusuarios pueden vaciar el catálogo")
    await db.execute(text("TRUNCATE TABLE precio_historial, productos RESTART IDENTITY CASCADE"))
    logger.info("vaciar_catalogo: tablas productos y precio_historial vaciadas")
    return {"message": "Catálogo vaciado correctamente"}


@router.get("/buscar-vivo")
async def buscar_vivo(
    q: str = Query(..., min_length=2, description="Término de búsqueda"),
    _: User = Depends(require_permission("precios.search")),
):
    """Búsqueda EN VIVO de un producto — no usa la base de datos, golpea las
    APIs de Ta-Ta, El Dorado, GDU, FarmaShop y Botiga en paralelo."""
    import asyncio
    loop = asyncio.get_running_loop()
    try:
        from app.services.scraper.live_search import buscar_todas
        resultados = await asyncio.wait_for(
            loop.run_in_executor(None, buscar_todas, q),
            timeout=90.0,
        )
    except asyncio.TimeoutError:
        raise HTTPException(status_code=504, detail="La búsqueda tardó demasiado. Probá con menos palabras.")
    except Exception as exc:
        logger.error("buscar_vivo: error para '%s': %s", q, exc, exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno en búsqueda en vivo")

    items = []
    for records in resultados.values():
        for r in records:
            items.append({
                "tienda":          r.tienda,
                "nombre":          r.nombre,
                "precio":          r.precio,
                "precio_lista":    r.precio_lista,
                "sku":             r.sku,
                "barcode":         r.barcode,
                "marca":           r.marca,
                "url":             r.url,
                "sucursal_id":     r.sucursal_id,
                "sucursal_nombre": r.sucursal_nombre,
            })

    return {"query": q, "total": len(items), "items": items}


def _resolver_barcode(barcode: str) -> str | None:
    """Consulta Open Food Facts para obtener el nombre de un producto por EAN.
    Retorna el nombre del producto o None si no se encuentra."""
    import requests as _req
    try:
        r = _req.get(
            f"https://world.openfoodfacts.org/api/v0/product/{barcode}.json",
            headers={"User-Agent": "MarketingTienda/1.0"},
            timeout=5,
        )
        if r.status_code == 200:
            data = r.json()
            if data.get("status") == 1:
                product = data.get("product", {})
                return (
                    product.get("product_name_es")
                    or product.get("product_name")
                    or product.get("generic_name_es")
                    or product.get("generic_name")
                )
    except Exception:
        pass
    return None


@router.get("/buscar-vivo-stream")
async def buscar_vivo_stream(
    q: str = Query(..., min_length=2, description="Término de búsqueda"),
    _: User = Depends(require_permission("precios.search")),
):
    """Búsqueda EN VIVO con SSE — devuelve resultados cadena por cadena en cuanto
    cada una termina. Evita el timeout de 30s de Render free tier porque los headers
    HTTP (incluyendo CORS) se envían con el primer byte, antes de que cualquier
    cadena termine."""
    import asyncio, json, threading
    from app.services.scraper.live_search import buscar_todas_streaming, _DATA_DIR

    # Si el término es puramente numérico, resolver barcode → nombre via Open Food Facts
    search_term = q.strip()
    if search_term.isdigit():
        nombre = await asyncio.get_event_loop().run_in_executor(
            None, _resolver_barcode, search_term
        )
        if nombre:
            search_term = nombre
        else:
            async def _not_found():
                payload = json.dumps({
                    "done": True,
                    "error": f"Código {search_term} no encontrado. Probá buscar por nombre del producto.",
                })
                yield f"data: {payload}\n\n"
            return StreamingResponse(_not_found(), media_type="text/event-stream")

    loop = asyncio.get_event_loop()
    queue: asyncio.Queue = asyncio.Queue()

    def _run_search():
        try:
            for cadena, records, error in buscar_todas_streaming(search_term, _DATA_DIR):
                try:
                    items = [
                        {
                            "tienda":          r.tienda,
                            "nombre":          r.nombre,
                            "precio":          r.precio,
                            "precio_lista":    r.precio_lista,
                            "sku":             r.sku,
                            "barcode":         r.barcode,
                            "marca":           r.marca,
                            "url":             r.url,
                            "sucursal_id":     r.sucursal_id,
                            "sucursal_nombre": r.sucursal_nombre,
                        }
                        for r in records
                        if r.nombre and r.precio
                    ]
                    payload: dict = {"cadena": cadena, "items": items}
                    if error:
                        payload["error"] = error
                    loop.call_soon_threadsafe(queue.put_nowait, json.dumps(payload))
                    logger.info("buscar_vivo_stream: %s OK — %d items para '%s'", cadena, len(items), q)
                except Exception as chain_exc:
                    logger.error("buscar_vivo_stream: error serializando %s — %s", cadena, chain_exc, exc_info=True)
                    fallback = json.dumps({"cadena": cadena, "items": [], "error": f"Error interno: {chain_exc}"})
                    loop.call_soon_threadsafe(queue.put_nowait, fallback)
        except Exception as exc:
            logger.error("buscar_vivo_stream: error iterando cadenas para '%s' — %s", q, exc, exc_info=True)
        finally:
            loop.call_soon_threadsafe(queue.put_nowait, None)  # sentinel

    threading.Thread(target=_run_search, daemon=True).start()

    async def generate():
        # Heartbeat every 5s — prevents Render/proxy from closing the connection
        # during the 20-30s gap while El Dorado / Ta-Ta are resolving timeouts.
        deadline = loop.time() + 120.0
        while True:
            remaining = deadline - loop.time()
            if remaining <= 0:
                logger.error("buscar_vivo_stream: timeout total para '%s'", q)
                yield 'data: {"done":true,"error":"timeout"}\n\n'
                break
            try:
                msg = await asyncio.wait_for(queue.get(), timeout=min(5.0, remaining))
            except asyncio.TimeoutError:
                yield ": keep-alive\n\n"  # SSE comment — browsers ignore, proxies see data
                continue
            if msg is None:
                yield 'data: {"done":true}\n\n'
                break
            yield f"data: {msg}\n\n"

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@router.get("/{producto_id}", response_model=ProductoOut)
async def obtener_precio(
    producto_id: int,
    _: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Producto).where(Producto.id == producto_id))
    prod = result.scalar_one_or_none()
    if not prod:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    return ProductoOut.model_validate(prod)
