"""Informe de produccion de cenefas.

Responde una pregunta concreta: cuantas cenefas se hicieron, cuantas salieron
correctas, y cuanto vale eso. Cada cenefa disenada tiene un costo (49 pesos
por defecto, configurable desde la pantalla) y ese es el criterio con el que
se valoriza el trabajo.

Cada corrida cuenta por separado, aunque sea el mismo listado reprocesado: es
trabajo pedido a la herramienta y se paga igual.

Solo entran los jobs CONFIRMADOS (status "done"). Un job en "preview" es una
previsualizacion que nadie confirmo, y uno en "error" no produjo nada.

De donde salen los numeros
--------------------------
Del `validation_report.summary` que el propio motor dejo al validar, no de un
recuento hecho ahora: es lo que se vio en pantalla en su momento, y eso es lo
que hay que poder demostrar.

    total            cenefas de esa corrida
    correct          las que no tienen ningun problema
    with_warnings    salieron, pero con algo para revisar (ej. descripcion larga)
    critical_errors  no se pudieron armar (ej. sin precio)

Los jobs viejos que no tienen summary caen a row_count / error_count, que son
columnas propias de la tabla y existen desde el principio.

Que se valoriza y que no
------------------------
No todo lo que paso por el motor es trabajo facturable, asi que el informe
separa en tres y solo el primero lleva plata:

    cobrable        mundos marcados `cobrable` (Rompe Precios, Parrilla y
                    Vinos, Mega Rompe Precios).
    sin costo       mundos con `cobrable=false`: Redexpres y el mundo de
                    pruebas. Se muestran con su volumen, valorizados en 0.
    sin clasificar  corridas sin `categoria` -- las de junio al 23/08, de antes
                    de que el job guardara el mundo. NO se valorizan: no se
                    sabe cuales fueron trabajo real y cuales pruebas o
                    reproceso, y valorizarlas es lo que inflaba el total.

Aparte va la produccion DECLARADA (`cenefa_destinos.cenefas_previas`): lo que
se hizo antes de que hubiera registro y solo se puede afirmar, no demostrar.
Suma al total cobrable pero viaja en su propio renglon y etiquetada, para que
se vea de un lado lo respaldado corrida por corrida y del otro lo declarado.
"""
from __future__ import annotations

from datetime import date, datetime
from typing import Any

from sqlalchemy import Integer, func, select, true
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.cenefa_destino import CenefaDestino
from app.models.cenefa_job import CenefaJob

# Costo por cenefa disenada, en pesos. Es solo el valor por defecto: la
# pantalla lo deja cambiar y el numero viaja en cada consulta, para que
# actualizarlo no dependa de un deploy.
COSTO_POR_CENEFA = 49.0


def _num(campo: str):
    """Un entero del summary del reporte, o NULL si ese job no lo tiene."""
    return (
        CenefaJob.validation_report[("summary", campo)]
        .astext.cast(Integer)
    )


def _metricas():
    """Las cuatro cifras de un job, con respaldo para los jobs viejos.

    Los primeros jobs no guardaban `validation_report.summary`; para esos se
    usan row_count y error_count, que son columnas de la tabla y estan desde
    el principio. Sin este respaldo, junio y julio aparecerian en cero.
    """
    total = func.coalesce(_num("total"), CenefaJob.row_count, 0)
    criticos = func.coalesce(_num("critical_errors"), CenefaJob.error_count, 0)
    correctas = func.coalesce(_num("correct"), CenefaJob.row_count - CenefaJob.error_count, 0)
    avisos = func.coalesce(_num("with_warnings"), 0)
    return total, correctas, avisos, criticos


def _filtros(desde: date | None, hasta: date | None, template: str | None):
    cond = [CenefaJob.status == "done"]
    if desde:
        cond.append(CenefaJob.created_at >= datetime.combine(desde, datetime.min.time()))
    if hasta:
        cond.append(CenefaJob.created_at < datetime.combine(hasta, datetime.max.time()))
    if template:
        cond.append(CenefaJob.template_nombre == template)
    return cond


async def _reales_por_categoria(db: AsyncSession, cond: list) -> dict[str, int]:
    """Cenefas reales por mundo: filas del listado x formatos distintos pedidos,
    sin contar de nuevo el reproceso.

    La corrida no es la unidad de trabajo -- el listado si. Un mismo Excel se
    reprocesa hasta que sale bien (Carniceria necesito 4 salidas y tuvo 25
    intentos en agosto), y el informe bruto paga cada intento por separado.
    Aca se agrupa por listado -- `excel_nombre` + `row_count`, NO solo el
    nombre: "convertidor_cenefas.xlsx" es el nombre generico que deja el
    Convertidor y se repite para listados distintos, lo unico que los separa
    es cuantas filas trae cada uno -- y dentro de cada listado se cuentan las
    plantillas (=formatos) distintas que efectivamente se generaron.
    Confirmado con Ivan el 2026-08-31 sobre Mega Rompe Precios.

    No usa la agrupacion de `agrupar_intentos` (excel + plantilla) porque esa
    clave puede juntar dos listados reales distintos que por casualidad usaron
    el mismo nombre generico Y la misma plantilla en fechas distintas, y se
    quedaria solo con el ultimo, perdiendo el otro. Por eso `row_count` entra
    en la clave.

    Solo cubre corridas con `excel_nombre` guardado (desde el 23/08/2026): las
    anteriores ya quedan afuera de lo cobrable por "sin clasificar", asi que
    no hace falta un respaldo para esas.
    """
    sub = (
        select(
            CenefaJob.categoria,
            CenefaJob.excel_nombre,
            CenefaJob.row_count,
            func.count(func.distinct(CenefaJob.template_nombre)).label("n_formatos"),
        )
        .where(*cond, CenefaJob.excel_nombre.isnot(None), CenefaJob.row_count.isnot(None))
        .group_by(CenefaJob.categoria, CenefaJob.excel_nombre, CenefaJob.row_count)
    ).subquery()

    filas = (await db.execute(
        select(
            sub.c.categoria,
            func.coalesce(func.sum(sub.c.row_count * sub.c.n_formatos), 0).label("reales"),
        ).group_by(sub.c.categoria)
    )).all()
    return {(r.categoria or ""): int(r.reales) for r in filas}


async def _reales_por_mes(db: AsyncSession, cond: list) -> dict[str, dict[str, int]]:
    """Cenefas reales por mes, con la misma regla que _reales_por_categoria:
    filas del listado x formatos distintos, sin contar de nuevo el reproceso.

    Devuelve {mes: {"reales": todas, "cobrables": solo mundos cobrables}}:
    el volumen mensual se muestra entero (Redexpres y pruebas incluidas),
    pero valorizar solo corresponde a lo cobrable -- igual que en el cuadro
    por mundo. Sin esta separacion, el informe valorizaba en agosto las
    reales de Redexpres como si se facturaran.

    Un listado reprocesado a caballo de dos meses cuenta en cada mes con los
    formatos que se le generaron ese mes: no hay una atribucion mejor sin
    inventar a que mes "pertenece" el listado."""
    mes = func.to_char(func.date_trunc("month", CenefaJob.created_at), "YYYY-MM")

    async def _sumar(solo_cobrables: bool) -> dict[str, int]:
        # La categoria ENTRA en la clave, igual que en _reales_por_categoria:
        # dos listados distintos pueden compartir nombre generico Y cantidad
        # de filas en mundos distintos (paso con convertidor_cenefas.xlsx de
        # 43 filas en Rompe Precios y en Pruebas), y sin la categoria el mes
        # los fusionaba y contaba 43 en vez de 86.
        q = select(
            mes.label("mes"),
            CenefaJob.categoria,
            CenefaJob.excel_nombre,
            CenefaJob.row_count,
            func.count(func.distinct(CenefaJob.template_nombre)).label("n_formatos"),
        )
        if solo_cobrables:
            # Mismo criterio que el cuadro por mundo: con mundo asignado y
            # cobrable (un mundo borrado cuenta cobrable, que es el default).
            q = (q.select_from(CenefaJob)
                 .outerjoin(CenefaDestino, CenefaDestino.slug == CenefaJob.categoria)
                 .where(CenefaJob.categoria.isnot(None),
                        func.coalesce(CenefaDestino.cobrable, true())))
        sub = (
            q.where(*cond, CenefaJob.excel_nombre.isnot(None),
                    CenefaJob.row_count.isnot(None))
            .group_by(mes, CenefaJob.categoria, CenefaJob.excel_nombre,
                      CenefaJob.row_count)
        ).subquery()
        filas = (await db.execute(
            select(
                sub.c.mes,
                func.coalesce(func.sum(sub.c.row_count * sub.c.n_formatos), 0).label("reales"),
            ).group_by(sub.c.mes)
        )).all()
        return {r.mes: int(r.reales) for r in filas}

    todas = await _sumar(solo_cobrables=False)
    cobrables = await _sumar(solo_cobrables=True)
    return {m: {"reales": n, "cobrables": cobrables.get(m, 0)}
            for m, n in todas.items()}


async def resumen(
    db: AsyncSession,
    desde: date | None = None,
    hasta: date | None = None,
    template: str | None = None,
    costo: float = COSTO_POR_CENEFA,
) -> dict[str, Any]:
    """Totales, apertura por mes y por plantilla."""
    total, correctas, avisos, criticos = _metricas()
    cond = _filtros(desde, hasta, template)

    # Lo verificado a mano se suma aparte: es la cifra que se puede defender
    # sin depender de que la validacion automatica no haya visto un problema.
    verif_corridas = func.count().filter(CenefaJob.verificado.is_(True))
    verif_cenefas = func.coalesce(
        func.sum(total).filter(CenefaJob.verificado.is_(True)), 0)

    fila = (await db.execute(
        select(
            func.count().label("corridas"),
            func.coalesce(func.sum(total), 0).label("cenefas"),
            func.coalesce(func.sum(correctas), 0).label("correctas"),
            func.coalesce(func.sum(avisos), 0).label("avisos"),
            func.coalesce(func.sum(criticos), 0).label("criticos"),
            verif_corridas.label("verif_corridas"),
            verif_cenefas.label("verif_cenefas"),
            func.min(CenefaJob.created_at).label("desde"),
            func.max(CenefaJob.created_at).label("hasta"),
        ).where(*cond)
    )).one()

    mes = func.to_char(func.date_trunc("month", CenefaJob.created_at), "YYYY-MM")
    por_mes = (await db.execute(
        select(
            mes.label("mes"),
            func.count().label("corridas"),
            func.coalesce(func.sum(total), 0).label("cenefas"),
            func.coalesce(func.sum(correctas), 0).label("correctas"),
            func.coalesce(func.sum(avisos), 0).label("avisos"),
            func.coalesce(func.sum(criticos), 0).label("criticos"),
            verif_corridas.label("verif_corridas"),
            verif_cenefas.label("verif_cenefas"),
        ).where(*cond).group_by(mes).order_by(mes)
    )).all()

    por_plantilla = (await db.execute(
        select(
            func.coalesce(CenefaJob.template_nombre, "").label("plantilla"),
            func.count().label("corridas"),
            func.coalesce(func.sum(total), 0).label("cenefas"),
            func.coalesce(func.sum(correctas), 0).label("correctas"),
            verif_corridas.label("verif_corridas"),
            verif_cenefas.label("verif_cenefas"),
        )
        .where(*cond)
        .group_by(CenefaJob.template_nombre)
        .order_by(func.coalesce(func.sum(total), 0).desc())
    )).all()

    # Apertura por mundo. El mundo se lee de la columna del job y NO de la
    # plantilla: la plantilla puede haberse borrado (FK ON DELETE SET NULL) y
    # ahi se perdia la atribucion. Un mundo borrado deja corridas con categoria
    # pero sin fila en cenefa_destinos: se asume cobrable, que es el default, y
    # se muestra con su slug como nombre.
    por_mundo = (await db.execute(
        select(
            func.coalesce(CenefaJob.categoria, "").label("mundo"),
            func.coalesce(CenefaDestino.nombre, "").label("nombre"),
            func.coalesce(CenefaDestino.cobrable, true()).label("cobrable"),
            func.count().label("corridas"),
            func.coalesce(func.sum(total), 0).label("cenefas"),
            func.coalesce(func.sum(correctas), 0).label("correctas"),
            func.coalesce(func.sum(avisos), 0).label("avisos"),
            func.coalesce(func.sum(criticos), 0).label("criticos"),
            verif_corridas.label("verif_corridas"),
            verif_cenefas.label("verif_cenefas"),
        )
        .select_from(CenefaJob)
        .outerjoin(CenefaDestino, CenefaDestino.slug == CenefaJob.categoria)
        .where(*cond)
        .group_by(CenefaJob.categoria, CenefaDestino.nombre, CenefaDestino.cobrable)
        .order_by(func.coalesce(func.sum(total), 0).desc())
    )).all()

    reales_por_categoria = await _reales_por_categoria(db, cond)
    reales_por_mes = await _reales_por_mes(db, cond)

    # Produccion declarada: una cifra fija por mundo, sin corridas detras. Solo
    # se incluye en la vista sin filtros -- acotada a un mes o a una plantilla
    # no significa nada, porque no tiene fecha ni plantilla que filtrar.
    sin_filtros = desde is None and hasta is None and not template
    declaradas = (await db.execute(
        select(
            CenefaDestino.slug, CenefaDestino.nombre, CenefaDestino.cobrable,
            CenefaDestino.cenefas_previas, CenefaDestino.cenefas_previas_nota,
        )
        .where(CenefaDestino.cenefas_previas > 0)
        .order_by(CenefaDestino.orden, CenefaDestino.slug)
    )).all() if sin_filtros else []

    _CAMPOS = ("corridas", "cenefas", "correctas", "avisos", "criticos",
               "verif_corridas", "verif_cenefas")

    def cifras(r) -> dict[str, int]:
        """Los numeros de una fila, con 0 para lo que esa consulta no trajo."""
        if isinstance(r, dict):
            return r
        return {k: (getattr(r, k, 0) or 0) for k in _CAMPOS}

    def bloque(r, valorizar: bool = True, reales: int | None = None) -> dict[str, Any]:
        """Un bloque de cifras. `valorizar=False` lo deja en cero pesos sin
        esconder el volumen: es lo que se hace con los mundos sin costo y con
        lo que no se puede atribuir.

        `reales` es lo que efectivamente hacia falta -- filas del listado x
        formatos distintos, sin contar de nuevo el reproceso. Se pasa aparte
        en vez de salir de `c` porque no es una columna mas: sale de agrupar
        por listado, no de sumar filas."""
        c = cifras(r)
        d = {
            "corridas":  c["corridas"],
            "cenefas":   c["cenefas"],
            "correctas": c["correctas"],
            "avisos":    c["avisos"],
            "criticos":  c["criticos"],
            "verificadas_corridas": c["verif_corridas"],
            "verificadas":          c["verif_cenefas"],
            "costo":             round(c["cenefas"] * costo, 2) if valorizar else 0.0,
            "costo_correctas":   round(c["correctas"] * costo, 2) if valorizar else 0.0,
            "costo_verificadas": round(c["verif_cenefas"] * costo, 2) if valorizar else 0.0,
        }
        if reales is not None:
            d["cenefas_reales"] = reales
            d["costo_real"] = round(reales * costo, 2) if valorizar else 0.0
        return d

    def sumar(filas) -> dict[str, int]:
        return {k: sum(cifras(r)[k] for r in filas) for k in _CAMPOS}

    # Tres grupos excluyentes. "Sin clasificar" es categoria NULL: las corridas
    # anteriores a que el job guardara el mundo.
    g_cobrable  = [r for r in por_mundo if r.mundo and r.cobrable]
    g_sin_costo = [r for r in por_mundo if r.mundo and not r.cobrable]
    g_sin_clas  = [r for r in por_mundo if not r.mundo]

    medido_cobrable    = sumar(g_cobrable)
    declarado_cobrable = sum(r.cenefas_previas for r in declaradas if r.cobrable)
    cenefas_cobrables  = medido_cobrable["cenefas"] + declarado_cobrable

    # Lo declarado ya es una cifra real (no tiene corridas ni reproceso
    # detras), asi que se suma directo -- lo unico que hay que desagregar del
    # reproceso es lo medido.
    reales_cobrable        = sum(reales_por_categoria.get(r.mundo, 0) for r in g_cobrable)
    cenefas_reales_cobrable = reales_cobrable + declarado_cobrable

    return {
        "costo_unitario": costo,
        # `total` sigue siendo TODO lo medido, valorizado como siempre: es la
        # cifra bruta, la que responde "cuanto paso por el motor".
        "total": {
            **bloque(fila),
            "desde": fila.desde.isoformat() if fila.desde else None,
            "hasta": fila.hasta.isoformat() if fila.hasta else None,
        },
        "por_mes": [
            {
                "mes": r.mes,
                **bloque(r, reales=(reales_por_mes.get(r.mes) or {}).get("reales")),
                # La plata del mes sale SOLO de lo cobrable; el volumen
                # (cenefas_reales) si muestra todo, como el cuadro por mundo.
                **({"costo_real": round(
                        (reales_por_mes.get(r.mes) or {}).get("cobrables", 0) * costo, 2)}
                   if r.mes in reales_por_mes else {}),
            }
            for r in por_mes
        ],
        "por_plantilla": [
            # Hasta el 23/08/2026 no se guardaba el nombre de la plantilla en el
            # job, asi que la mitad del historial no puede atribuirse. Se muestra
            # como un grupo aparte en vez de esconderlo.
            {"plantilla": r.plantilla or "(sin registrar)", **bloque(r)}
            for r in por_plantilla
        ],
        "por_mundo": [
            {
                "mundo":    r.mundo or "",
                "nombre":   r.nombre or r.mundo or "(sin clasificar)",
                "cobrable": bool(r.mundo) and bool(r.cobrable),
                **bloque(r, valorizar=bool(r.mundo) and bool(r.cobrable),
                         reales=reales_por_categoria.get(r.mundo or "")),
            }
            for r in por_mundo
        ],
        "declaradas": [
            {
                "mundo":    r.slug,
                "nombre":   r.nombre,
                "cobrable": bool(r.cobrable),
                "cenefas":  r.cenefas_previas,
                "nota":     r.cenefas_previas_nota,
                "costo":    round(r.cenefas_previas * costo, 2) if r.cobrable else 0.0,
            }
            for r in declaradas
        ],
        # Lo que de verdad se factura: mundos cobrables, medido + declarado.
        "cobrable": {
            **bloque(medido_cobrable, reales=reales_cobrable),
            "declaradas":      declarado_cobrable,
            "cenefas_totales": cenefas_cobrables,
            "costo_total":     round(cenefas_cobrables * costo, 2),
            # Lo que hacia falta de verdad -- listado x formatos distintos,
            # sin pagar de nuevo el reproceso. Medido con esa regla + lo
            # declarado (que ya es real, no tiene corridas detras).
            "cenefas_reales_totales": cenefas_reales_cobrable,
            "costo_real_total":       round(cenefas_reales_cobrable * costo, 2),
        },
        "sin_costo":      bloque(sumar(g_sin_costo), valorizar=False),
        "sin_clasificar": bloque(sumar(g_sin_clas),  valorizar=False),
    }


async def detalle(
    db: AsyncSession,
    desde: date | None = None,
    hasta: date | None = None,
    template: str | None = None,
    costo: float = COSTO_POR_CENEFA,
    limite: int = 500,
) -> list[dict[str, Any]]:
    """Una fila por corrida, de la mas nueva a la mas vieja."""
    total, correctas, avisos, criticos = _metricas()
    filas = (await db.execute(
        select(
            CenefaJob.id, CenefaJob.created_at, CenefaJob.completed_at,
            CenefaJob.format, CenefaJob.template_nombre, CenefaJob.excel_nombre,
            CenefaJob.created_by,
            CenefaJob.verificado, CenefaJob.verificado_at,
            total.label("cenefas"), correctas.label("correctas"),
            avisos.label("avisos"), criticos.label("criticos"),
        )
        .where(*_filtros(desde, hasta, template))
        .order_by(CenefaJob.created_at.desc())
        .limit(limite)
    )).all()

    return [
        {
            "id":        str(r.id),
            "fecha":     r.created_at.isoformat() if r.created_at else None,
            "formato":   r.format,
            "plantilla": r.template_nombre or "(sin registrar)",
            "excel":     r.excel_nombre or "(sin registrar)",
            "usuario_id": r.created_by,
            "cenefas":   r.cenefas or 0,
            "correctas": r.correctas or 0,
            "avisos":    r.avisos or 0,
            "criticos":  r.criticos or 0,
            "verificado": bool(r.verificado),
            "verificado_at": r.verificado_at.isoformat() if r.verificado_at else None,
            "costo":     round((r.cenefas or 0) * costo, 2),
        }
        for r in filas
    ]


async def plantillas_del_historial(db: AsyncSession) -> list[str]:
    """Nombres de plantilla que aparecen en el historial, para el filtro."""
    filas = (await db.execute(
        select(CenefaJob.template_nombre)
        .where(CenefaJob.status == "done", CenefaJob.template_nombre.isnot(None))
        .distinct()
        .order_by(CenefaJob.template_nombre)
    )).scalars().all()
    return list(filas)


# ---------------------------------------------------------------------------
# Excel del informe
# ---------------------------------------------------------------------------

def a_excel(resumen_: dict[str, Any], filas: list[dict[str, Any]]) -> bytes:
    """El informe en Excel: una hoja de resumen y otra con el detalle.

    El resumen es lo que se muestra o se imprime; el detalle esta para poder
    respaldar cualquier cifra del resumen corrida por corrida, que es de lo
    que se trata poder demostrar lo que se hizo.
    """
    import io

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    costo = resumen_["costo_unitario"]
    wb = openpyxl.Workbook()

    titulo_font = Font(bold=True, size=14)
    head_fill = PatternFill("solid", fgColor="1E3A5F")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    dinero = '#,##0'

    def encabezar(ws, fila, nombres):
        for c, n in enumerate(nombres, 1):
            cell = ws.cell(row=fila, column=c, value=n)
            cell.fill, cell.font = head_fill, head_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Resumen ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "Informe de produccion de cenefas"
    ws["A1"].font = titulo_font
    t = resumen_["total"]
    periodo = f'{(t["desde"] or "")[:10]} a {(t["hasta"] or "")[:10]}' if t["desde"] else "sin datos"
    ws["A2"] = f'Periodo: {periodo}    Costo por cenefa: ${costo:g}'

    # Sin "Correctas" (las reales YA son las que salieron bien -- Ivan,
    # 01/09) y sin "Valor total": valorizar el bruto global (Redexpres,
    # pruebas y sin clasificar incluidos) es justo el total inflado que este
    # informe vino a matar. La plata vive solo en "Que se factura".
    ws["A4"] = "Total"
    ws["A4"].font = Font(bold=True)
    encabezar(ws, 5, ["Corridas", "Cenefas (bruto, referencia)", "Con avisos",
                      "No se pudieron armar",
                      "Corridas verificadas", "Cenefas verificadas", "Valor verificado"])
    ws.cell(row=6, column=1, value=t["corridas"])
    ws.cell(row=6, column=2, value=t["cenefas"])
    ws.cell(row=6, column=3, value=t["avisos"])
    ws.cell(row=6, column=4, value=t["criticos"])
    ws.cell(row=6, column=5, value=t["verificadas_corridas"])
    ws.cell(row=6, column=6, value=t["verificadas"])
    ws.cell(row=6, column=7, value=t["costo_verificadas"]).number_format = dinero

    # ── Que se factura y que no ────────────────────────────────────────────
    # El "Total" de arriba es la cifra bruta: todo lo que paso por el motor.
    # Este bloque es el que se puede presentar, porque separa el trabajo
    # cobrable de los mundos sin costo y de lo que no se puede atribuir.
    cob = resumen_.get("cobrable") or {}
    sinc = resumen_.get("sin_costo") or {}
    sincl = resumen_.get("sin_clasificar") or {}
    declaradas = resumen_.get("declaradas") or []

    # Mismo armado que el PDF semanal: componentes de lo real, TOTAL REAL en
    # negrita, y el bruto como referencia. "Cobrable" se reserva para lo
    # real -- llamarle cobrable al bruto invitaba a tomar el numero inflado.
    ws["A8"] = "Que se factura"
    ws["A8"].font = Font(bold=True)
    encabezar(ws, 9, ["Concepto", "Cenefas", "Valor", "Detalle"])
    fila = 10
    mundos_medidos = " + ".join(
        f'{r["nombre"]} ({r["cenefas_reales"]})'
        for r in resumen_["por_mundo"]
        if r["cobrable"] and r.get("cenefas_reales")
    ) or "Mundos cobrables"
    presentacion = [
        ("Medido (reales)", cob.get("cenefas_reales", 0), cob.get("costo_real", 0),
         f"{mundos_medidos} - listado x formatos, sin contar de nuevo el reproceso"),
    ]
    for d in declaradas:
        if not d["cobrable"]:
            continue
        presentacion.append((
            f'Declarado - {d["nombre"]}', d["cenefas"], d["costo"], d["nota"],
        ))
    presentacion += [
        ("TOTAL REAL", cob.get("cenefas_reales_totales", 0),
         cob.get("costo_real_total", 0),
         "Medido + declarado - el numero que se factura"),
        ("Referencia: bruto medido", cob.get("cenefas", 0), cob.get("costo", 0),
         "Cada reproceso cuenta de nuevo; no se factura"),
        ("Sin costo", sinc.get("cenefas", 0), 0,
         "Mundos marcados sin costo (Redexpres, pruebas)"),
        ("Sin clasificar", sincl.get("cenefas", 0), 0,
         "Pre-23/08, sin mundo registrado; no se valoriza"),
    ]
    for concepto, cenefas_, valor_, detalle_ in presentacion:
        ws.cell(row=fila, column=1, value=concepto)
        if concepto == "TOTAL REAL":
            for c in (1, 2, 3):
                ws.cell(row=fila, column=c).font = Font(bold=True)
        ws.cell(row=fila, column=2, value=cenefas_)
        ws.cell(row=fila, column=3, value=valor_).number_format = dinero
        ws.cell(row=fila, column=4, value=detalle_)
        fila += 1

    fila += 2
    # Brutas + Reales + Valor real, como el PDF y la web. Por plantilla no
    # tiene reales (cada fila ES un formato): su bruto va rotulado como
    # referencia. En Por mundo, lo declarado entra como fila propia para que
    # Parrilla y Vinos no desaparezca por no tener corridas.
    for titulo, clave, etiqueta in (("Por mundo", "por_mundo", "nombre"),
                                    ("Por mes", "por_mes", "mes"),
                                    ("Por plantilla", "por_plantilla", "plantilla")):
        con_reales = clave != "por_plantilla"
        ws.cell(row=fila, column=1, value=titulo).font = Font(bold=True)
        rotulo = "Mundo" if clave == "por_mundo" else etiqueta.capitalize()
        encabezar(ws, fila + 1,
                  [rotulo, "Corridas", "Brutas", "Reales", "Valor real"]
                  if con_reales else
                  [rotulo, "Corridas", "Brutas", "Valor bruto (referencia)"])
        fila += 2
        for r in resumen_[clave]:
            ws.cell(row=fila, column=1, value=r[etiqueta])
            ws.cell(row=fila, column=2, value=r["corridas"])
            ws.cell(row=fila, column=3, value=r["cenefas"])
            if con_reales:
                ws.cell(row=fila, column=4, value=r.get("cenefas_reales"))
                ws.cell(row=fila, column=5,
                        value=r.get("costo_real") or None).number_format = dinero
            else:
                ws.cell(row=fila, column=4, value=r["costo"]).number_format = dinero
            fila += 1
        if clave == "por_mundo":
            for d in declaradas:
                if not d["cobrable"]:
                    continue
                ws.cell(row=fila, column=1, value=f'{d["nombre"]} (declarado)')
                ws.cell(row=fila, column=4, value=d["cenefas"])
                ws.cell(row=fila, column=5, value=d["costo"]).number_format = dinero
                fila += 1
        fila += 2

    for col, ancho in zip("ABCDEFGHI", (34, 12, 12, 12, 20, 16, 20, 20, 18)):
        ws.column_dimensions[col].width = ancho

    # ── Detalle ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Detalle")
    # Sin "Correctas" (Ivan, 01/09) y "Valor bruto (ref.)" en vez de "Valor":
    # cada reproceso figura con su bruto x costo como referencia de auditoria,
    # pero el reproceso no se factura.
    cols = ["Fecha", "Plantilla", "Excel", "Formato", "Cenefas",
            "Con avisos", "No se pudieron armar", "Valor bruto (ref.)", "Verificada"]
    encabezar(ws2, 1, cols)
    for i, r in enumerate(filas, 2):
        ws2.cell(row=i, column=1, value=(r["fecha"] or "")[:16].replace("T", " "))
        ws2.cell(row=i, column=2, value=r["plantilla"])
        ws2.cell(row=i, column=3, value=r["excel"])
        ws2.cell(row=i, column=4, value=r["formato"])
        ws2.cell(row=i, column=5, value=r["cenefas"])
        ws2.cell(row=i, column=6, value=r["avisos"])
        ws2.cell(row=i, column=7, value=r["criticos"])
        ws2.cell(row=i, column=8, value=r["costo"]).number_format = dinero
        ws2.cell(row=i, column=9, value="si" if r["verificado"] else "")
    for c, ancho in enumerate((17, 34, 34, 10, 10, 12, 21, 16, 12), 1):
        ws2.column_dimensions[get_column_letter(c)].width = ancho
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Historial de intentos
# ---------------------------------------------------------------------------

def agrupar_intentos(filas: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Agrupa las corridas por listado, para ver cuantos intentos llevo cada uno.

    Un mismo listado se reprocesa varias veces: se genera, se ve algo mal, se
    arregla y se vuelve a generar. En la lista plana eso son N corridas sueltas
    y no se entiende nada; agrupadas se lee de una que "este listado llevo 9
    intentos y el ultimo salio limpio".

    La clave es el Excel + la plantilla. Las corridas anteriores al 23/08/2026
    no guardaron esos nombres, asi que para esas se agrupa por formato y
    cantidad de carteles, que es lo unico que las distingue.

    Dentro de cada grupo los intentos van del mas viejo al mas nuevo, que es
    como se leen: el ultimo es el que quedo.
    """
    grupos: dict[tuple[str, str], dict[str, Any]] = {}
    for f in filas:
        excel = f["excel"] if f["excel"] != "(sin registrar)" else f'{f["cenefas"]} carteles'
        plantilla = f["plantilla"] if f["plantilla"] != "(sin registrar)" else (f["formato"] or "?")
        g = grupos.setdefault((excel, plantilla), {
            "excel": excel, "plantilla": plantilla,
            "sin_registrar": f["excel"] == "(sin registrar)",
            "intentos": [],
        })
        g["intentos"].append(f)

    salida = []
    for g in grupos.values():
        # Del mas viejo al mas nuevo: el ultimo es el que quedo.
        g["intentos"].sort(key=lambda x: x["fecha"] or "")
        ultimo = g["intentos"][-1]
        salida.append({
            "excel": g["excel"],
            "plantilla": g["plantilla"],
            "sin_registrar": g["sin_registrar"],
            "intentos": len(g["intentos"]),
            "primera": g["intentos"][0]["fecha"],
            "ultima": ultimo["fecha"],
            # Lo que vale es el ultimo intento: los anteriores se descartaron.
            "cenefas": ultimo["cenefas"],
            "correctas": ultimo["correctas"],
            "criticos": ultimo["criticos"],
            "verificado": ultimo["verificado"],
            "costo": ultimo["costo"],
            "detalle": g["intentos"],
        })
    # Primero los que mas intentos llevaron: son los que hay que mirar.
    salida.sort(key=lambda x: (-x["intentos"], x["ultima"] or ""), reverse=False)
    return salida


# ---------------------------------------------------------------------------
# PDF del informe — pensado para mandarse como reporte semanal
# ---------------------------------------------------------------------------

_MESES = ("enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
          "agosto", "setiembre", "octubre", "noviembre", "diciembre")


def _mes_legible(yyyymm: str) -> str:
    try:
        anio, mes = yyyymm.split("-")
        return f"{_MESES[int(mes) - 1]} {anio}"
    except (ValueError, IndexError):
        return yyyymm


def a_pdf(resumen_: dict[str, Any]) -> bytes:
    """El informe como reporte semanal: una o dos paginas para mandar, con lo
    real adelante y el respaldo (declarado, bruto, sin costo) detras.

    El Excel queda para auditar corrida por corrida; esto es lo que se
    presenta. Por eso el orden: primero el numero que se defiende (reales y
    su valor), despues de donde sale (que se factura), y recien al final los
    desgloses. Todo el texto es latin-1: Helvetica no tiene rayas largas ni
    flechas, y un caracter fuera de tabla revienta la descarga entera.
    """
    import io
    from datetime import datetime as _dt

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    AZUL = colors.HexColor("#1E3A5F")
    GRIS = colors.HexColor("#64748B")
    LINEA = colors.HexColor("#E2E8F0")
    FONDO = colors.HexColor("#F1F5F9")

    def n(x) -> str:
        return f"{int(round(x or 0)):,}".replace(",", ".")

    def pesos(x) -> str:
        return f"$ {n(x)}"

    st_titulo = ParagraphStyle("titulo", fontName="Helvetica-Bold",
                               fontSize=17, textColor=AZUL, leading=21)
    st_sub = ParagraphStyle("sub", fontName="Helvetica", fontSize=9,
                            textColor=GRIS, leading=13)
    st_seccion = ParagraphStyle("seccion", fontName="Helvetica-Bold",
                                fontSize=11, textColor=AZUL, leading=14,
                                spaceBefore=6)
    st_celda = ParagraphStyle("celda", fontName="Helvetica", fontSize=8,
                              leading=10)
    st_nota = ParagraphStyle("nota", fontName="Helvetica", fontSize=7,
                             textColor=GRIS, leading=9)
    st_kpi_num = ParagraphStyle("kpinum", fontName="Helvetica-Bold",
                                fontSize=16, textColor=AZUL, leading=19,
                                alignment=1)
    st_kpi_lbl = ParagraphStyle("kpilbl", fontName="Helvetica", fontSize=7.5,
                                textColor=GRIS, leading=10, alignment=1)
    st_pie = ParagraphStyle("pie", fontName="Helvetica", fontSize=7.5,
                            textColor=GRIS, leading=10)

    costo = resumen_["costo_unitario"]
    t = resumen_["total"]
    cob = resumen_.get("cobrable") or {}
    sinc = resumen_.get("sin_costo") or {}
    sincl = resumen_.get("sin_clasificar") or {}
    declaradas = resumen_.get("declaradas") or []

    periodo = (f'{(t["desde"] or "")[:10]} al {(t["hasta"] or "")[:10]}'
               if t.get("desde") else "sin datos")

    flujo: list[Any] = [
        Paragraph("Informe semanal de cenefas", st_titulo),
        Paragraph(
            f"Producci&oacute;n y valorizaci&oacute;n &middot; Per&iacute;odo: {periodo} "
            f"&middot; Costo por cenefa: ${costo:g} "
            f"&middot; Generado el {_dt.now().strftime('%d/%m/%Y')}",
            st_sub),
        Spacer(0, 6 * mm),
    ]

    # ── KPIs: el numero que se manda, primero ──────────────────────────────
    # Un "0 verificadas" gigante en el reporte que se presenta lee como que
    # nada salio bien, cuando en realidad nadie tildo el check todavia:
    # mientras no se use, ese lugar muestra el costo vigente.
    kpi_verif = ((n(t.get("verificadas", 0)), "VERIFICADAS A MANO")
                 if t.get("verificadas_corridas") else
                 (f"${costo:g}", "COSTO POR CENEFA"))
    kpis = [
        (n(cob.get("cenefas_reales_totales", 0)), "CENEFAS REALES"),
        (pesos(cob.get("costo_real_total", 0)), "VALOR REAL"),
        kpi_verif,
        (n(t.get("corridas", 0)), "CORRIDAS"),
    ]
    tabla_kpi = Table(
        [[Paragraph(v, st_kpi_num) for v, _ in kpis],
         [Paragraph(l, st_kpi_lbl) for _, l in kpis]],
        colWidths=[44 * mm] * 4,
    )
    tabla_kpi.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), FONDO),
        ("BOX", (0, 0), (-1, -1), 0.5, LINEA),
        ("LINEBEFORE", (1, 0), (-1, -1), 0.5, LINEA),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 8),
    ]))
    flujo += [tabla_kpi, Spacer(0, 7 * mm)]

    def tabla_seccion(titulo: str, cabecera: list[str], cuerpo: list[list],
                      anchos: list[float], resaltar: set[int] | None = None,
                      ultima_izquierda: bool = False):
        flujo.append(Paragraph(titulo, st_seccion))
        flujo.append(Spacer(0, 2 * mm))
        tabla = Table([cabecera] + cuerpo, colWidths=[a * mm for a in anchos],
                      repeatRows=1)
        estilo = [
            ("BACKGROUND", (0, 0), (-1, 0), AZUL),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, FONDO]),
            ("LINEBELOW", (0, 0), (-1, -1), 0.4, LINEA),
            ("TOPPADDING", (0, 0), (-1, -1), 3.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]
        if ultima_izquierda:
            estilo.append(("ALIGN", (-1, 0), (-1, -1), "LEFT"))
        for i in resaltar or set():
            estilo += [("FONTNAME", (0, i + 1), (-1, i + 1), "Helvetica-Bold"),
                       ("BACKGROUND", (0, i + 1), (-1, i + 1),
                        colors.HexColor("#DBEAFE"))]
        tabla.setStyle(TableStyle(estilo))
        # extend y no +=: += rebindea el nombre y rompe la clausura.
        flujo.extend([tabla, Spacer(0, 6 * mm)])

    # ── Que se factura ──────────────────────────────────────────────────────
    # El detalle del medido nombra los mundos de donde salen las cenefas
    # (Ivan lo pidio el 01/09: sin esto parecia que todo era Mega Rompe).
    mundos_medidos = [
        f'{r["nombre"]} ({n(r["cenefas_reales"])})'
        for r in resumen_["por_mundo"]
        if r["cobrable"] and r.get("cenefas_reales")
    ]
    detalle_medido = (
        f'{" + ".join(mundos_medidos)}: filas del listado x formatos '
        "generados, sin contar de nuevo el reproceso"
        if mundos_medidos else
        "Mundos cobrables: filas del listado x formatos generados, "
        "sin contar de nuevo el reproceso"
    )
    cuerpo: list[list] = [[
        Paragraph("Medido (reales)", st_celda),
        n(cob.get("cenefas_reales", 0)), pesos(cob.get("costo_real", 0)),
        Paragraph(detalle_medido, st_nota),
    ]]
    for d in declaradas:
        if not d["cobrable"]:
            continue
        cuerpo.append([
            Paragraph(f'Declarado &middot; {d["nombre"]}', st_celda),
            n(d["cenefas"]), pesos(d["costo"]),
            Paragraph(d["nota"] or "", st_nota),
        ])
    fila_total = len(cuerpo)
    cuerpo.append([
        Paragraph("TOTAL REAL", st_celda),
        n(cob.get("cenefas_reales_totales", 0)),
        pesos(cob.get("costo_real_total", 0)),
        Paragraph(f"Medido + declarado, a ${costo:g} por cenefa", st_nota),
    ])
    cuerpo += [
        [Paragraph("Referencia: bruto medido", st_celda),
         n(cob.get("cenefas", 0)), pesos(cob.get("costo", 0)),
         Paragraph("Todo lo que paso por el motor; cada reproceso cuenta "
                   "de nuevo", st_nota)],
        [Paragraph("Sin costo", st_celda), n(sinc.get("cenefas", 0)), "-",
         Paragraph("Mundos marcados sin costo (Redexpres, pruebas)", st_nota)],
        [Paragraph("Sin clasificar", st_celda), n(sincl.get("cenefas", 0)), "-",
         Paragraph("Corridas anteriores a que el job guardara el mundo",
                   st_nota)],
    ]
    tabla_seccion("Qu&eacute; se factura", ["Concepto", "Cenefas", "Valor", "Detalle"],
                  cuerpo, [42, 20, 26, 90], resaltar={fila_total},
                  ultima_izquierda=True)

    # ── Por mundo ───────────────────────────────────────────────────────────
    # Sin columna "Correctas": las reales YA son las que salieron bien (Ivan,
    # 01/09) -- la validacion automatica sobre el bruto al lado de las reales
    # solo confundia. Lo declarado entra como fila propia para que Parrilla y
    # Vinos no desaparezca del cuadro por no tener corridas.
    # "(sin clasificar)" al final y con leyenda propia: no es lo mismo que
    # sin costo (Redexpres no se cobra por decision; esto no tiene mundo) y
    # sus 25 mil brutas no pueden encabezar el cuadro.
    cuerpo = []
    con_mundo = [r for r in resumen_["por_mundo"] if r["mundo"]]
    sin_clas = [r for r in resumen_["por_mundo"] if not r["mundo"]]
    for r in con_mundo:
        nombre = r["nombre"] + ("" if r["cobrable"] else " (sin costo)")
        reales = n(r["cenefas_reales"]) if "cenefas_reales" in r else "-"
        valor = pesos(r["costo_real"]) if r.get("costo_real") else "-"
        cuerpo.append([Paragraph(nombre, st_celda), n(r["corridas"]),
                       n(r["cenefas"]), reales, valor])
    for d in declaradas:
        if not d["cobrable"]:
            continue
        cuerpo.append([Paragraph(f'{d["nombre"]} (declarado)', st_celda),
                       "-", "-", n(d["cenefas"]), pesos(d["costo"])])
    for r in sin_clas:
        reales = n(r["cenefas_reales"]) if "cenefas_reales" in r else "-"
        cuerpo.append([Paragraph(
            "(sin clasificar) - pre-23/08, sin mundo; no se valoriza",
            st_nota), n(r["corridas"]), n(r["cenefas"]), reales, "-"])
    tabla_seccion("Por mundo",
                  ["Mundo", "Corridas", "Brutas", "Reales", "Valor real"],
                  cuerpo, [66, 24, 24, 24, 32])

    # ── Por mes ─────────────────────────────────────────────────────────────
    cuerpo = []
    for r in resumen_["por_mes"]:
        reales = n(r["cenefas_reales"]) if "cenefas_reales" in r else "-"
        valor = pesos(r["costo_real"]) if r.get("costo_real") else "-"
        cuerpo.append([Paragraph(_mes_legible(r["mes"]), st_celda),
                       n(r["corridas"]), n(r["cenefas"]), reales, valor])
    tabla_seccion("Por mes",
                  ["Mes", "Corridas", "Brutas", "Reales", "Valor real"],
                  cuerpo, [66, 24, 24, 24, 32])

    flujo.append(Paragraph(
        "Reales = filas del listado x formatos distintos generados, sin pagar "
        "de nuevo el reproceso. Brutas = todo lo que paso por el motor. "
        "Declarado = produccion anterior al registro en la plataforma, con la "
        "nota que la respalda. El Valor real valoriza solo los mundos "
        "cobrables. El detalle corrida por corrida esta en el Excel del "
        "informe.", st_pie))

    buf = io.BytesIO()
    SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=16 * mm, rightMargin=16 * mm,
        topMargin=16 * mm, bottomMargin=14 * mm,
        title="Informe semanal de cenefas",
    ).build(flujo)
    return buf.getvalue()
