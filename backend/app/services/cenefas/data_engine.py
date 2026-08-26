"""Parseo de Excel/CSV — motor de datos para cenefas.

Este módulo ya NO calcula nada: no arma mecánicas de combo, no divide
precios por kilo, no infiere títulos. Todo eso vive ahora en el Convertidor
(``convertidor.py``), que deja el Excel con una columna por variable y el
valor final ya escrito. Acá solo se leen esas columnas y se las normaliza.

El cambio es a pedido explícito (08/2026): "que el convertidor sea el que
se encargue de dejar todo en columnas, para que acá simplemente nos
encarguemos de sustituir variables". Antes la misma lógica de combos vivía
duplicada en tres lugares con reglas ligeramente distintas, y un Excel
cargado a mano podía disparar sin querer reglas pensadas para el export de
gestión (caso real: una columna "OFERTADET" de Parrilla y Vinos activaba la
lógica de combos de Redexpres y pisaba precios que ya venían resueltos).
"""
import io
import re

import openpyxl

from app.services.cenefas.formatters import fmt_price, parse_price_raw
from app.services.cenefas.variables import (
    CANONICAL_VARS,
    DECIMAL_OF,
    INTERNAL_FIELDS,
    LEGAL_ALCOHOL,
    ORDEN_EXPORT,
    PRICE_VARS,
    categoria_es_alcohol,
    resolve,
)

CANONICAL_SET = frozenset(CANONICAL_VARS)

# Columnas que sirven para detectar la fila de encabezados: si una fila
# contiene al menos una, es la fila de headers.
_DETECTION_VARS = {"codigo", "descripcion", "precioRegular", "precioOferta"}

_RE_PRECIO_LIMPIO = re.compile(r"[\$\s]*\d[\d.,]*\s*")

# "1.299" / "1.234.567": puntos seguidos SIEMPRE de exactamente tres dígitos y
# sin coma decimal. Es separador de miles, no un punto decimal -- ningún
# precio tiene tres decimales.
#
# Hace falta porque el Convertidor escribe los precios ya formateados al
# estilo uruguayo ("1.299"), y ese mismo Excel se vuelve a leer acá al generar
# la cenefa. Sin esta regla, parse_price_raw() interpreta "1.299" como 1,299 y
# el cartel sale con "$1" en vez de "$1.299" (bug real de ida y vuelta, sin
# ningún síntoma hasta que un precio pasa los cuatro dígitos).
_RE_MILES_CON_PUNTO = re.compile(r"^\d{1,3}(?:\.\d{3})+$")

# "39 unidad", "31,2 unidad", "74,5 la unidad", "1.919,20 c/u": el precio con
# la unidad de venta escrita al lado. Gestión lo manda así en varios listados
# y hasta ahora sobrevivía al render: el cartel salía con "$39 unidad" en el
# cuadro del precio gigante y, peor, "31,2 unidad" ni siquiera se partía, así
# que el cuadro de decimales quedaba vacío en vez de mostrar ",20".
#
# Solo se limpia cuando el valor ARRANCA con el número y lo que sigue son
# letras: así "2x1", "3x99" y "SOLO X 25" siguen pasando tal cual, que es a
# propósito -- en una mecánica M x N el literal ocupa el cuadro del precio.
#
# El Convertidor ya limpiaba esto (_parse_price_or_none); esto lo lleva a la
# subida directa del Excel, que es por donde entran los listados de gestión
# que no pasan por el Convertidor.
_RE_PRECIO_CON_UNIDAD = re.compile(
    r"^\s*\$?\s*(\d[\d.,]*)\s+[a-zA-ZáéíóúñÁÉÍÓÚÑ/][a-zA-ZáéíóúñÁÉÍÓÚÑ/.\s]*$"
)


# ---------------------------------------------------------------------------
# Precio -> entero + decimal
# ---------------------------------------------------------------------------

def split_price(raw) -> tuple[str, str]:
    """Devuelve (entero, decimal) para un valor de precio crudo.

    - ``1234.5``      -> ("1.234", ",50")
    - ``"1.234,50"``  -> ("1.234", ",50")
    - ``899``         -> ("899", "")        redondo: sin decimal
    - ``""`` / None   -> ("", "")
    - ``"2x1"``       -> ("2x1", "")        no es un precio: pasa tal cual

    El decimal SIEMPRE lleva la coma adelante (decisión explícita) y queda
    vacío cuando el precio es redondo -- así el cuadro de decimales de la
    PPT no imprime un ",00" que el diseño no contempla.

    El caso "no es un precio" es intencional y necesario: en las mecánicas
    M x N la variable ``precioOferta`` deja de ser un número y pasa a ser el
    literal de la oferta ("2x1"), que ocupa el mismo cuadro grande donde
    iría el precio. Si acá se forzara a número, ese texto se perdería.
    """
    if raw is None:
        return "", ""
    if isinstance(raw, str) and not raw.strip():
        return "", ""

    if isinstance(raw, (int, float)):
        valor = float(raw)
    else:
        texto = str(raw).strip()
        # "39 unidad" -> "39": se saca la unidad de venta y sigue el camino
        # normal, así el decimal también se separa ("31,2 unidad" -> 31 + ",20").
        con_unidad = _RE_PRECIO_CON_UNIDAD.match(texto)
        if con_unidad:
            texto = con_unidad.group(1)
        # Solo se parsea como precio si el valor es un número de punta a
        # punta. `parse_price_raw` matchea el run de dígitos inicial e
        # ignora el resto en silencio ("4x350" -> 4.0), que acá convertiría
        # un M x N en un precio falso.
        if not _RE_PRECIO_LIMPIO.fullmatch(texto):
            return texto, ""
        limpio = texto.replace("$", "").strip()
        if _RE_MILES_CON_PUNTO.fullmatch(limpio):
            limpio = limpio.replace(".", "")
        valor = parse_price_raw(limpio)

    if valor <= 0:
        return ("" if isinstance(raw, (int, float)) else str(raw).strip()), ""

    formateado = fmt_price(valor)          # "1.234,50" | "899"
    if "," in formateado:
        entero, decimal = formateado.rsplit(",", 1)
        # Un decimal en cero es un precio redondo escrito raro (1234.00):
        # mismo tratamiento que un redondo, sin ",00" en el diseño.
        if decimal.strip("0") == "":
            return entero, ""
        return entero, "," + decimal
    return formateado, ""


def normalize_decimal(raw) -> str:
    """Normaliza una columna de decimales ya escrita por el Convertidor.

    Acepta ",50" / "50" / 0.5 y devuelve siempre ",50". Vacío si no hay nada
    o si el decimal es cero.
    """
    if raw is None:
        return ""

    # Excel puede guardar lo que la persona escribió como "0,50" en la
    # celda como el número 0.5. Sin este caso, el strip de no-dígitos de
    # abajo lo leería como ",05" -- cincuenta centavos impresos como cinco.
    if isinstance(raw, float) and 0 < raw < 1:
        return "," + f"{round(raw * 100):02d}"

    texto = str(raw).strip()
    if not texto:
        return ""
    solo_digitos = re.sub(r"[^\d]", "", texto)
    if not solo_digitos or solo_digitos.strip("0") == "":
        return ""
    return "," + solo_digitos[:2].ljust(2, "0")


# ---------------------------------------------------------------------------
# Procesamiento de fila
# ---------------------------------------------------------------------------

def process_row(
    row: tuple,
    h: dict,                    # var_name -> col_idx
    *,
    vigencia: str = "",
    legales: str = "",
    usar_legales: bool = False,
) -> dict:
    """Convierte una fila de Excel en el dict de variables del renderer."""
    result: dict[str, str] = {}

    def celda(var: str):
        idx = h.get(var)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    # -- Precios: entero + decimal ---------------------------------------
    for var in PRICE_VARS:
        decimal_var = DECIMAL_OF[var]
        entero, decimal_auto = split_price(celda(var))
        result[var] = entero

        # La columna de decimales explícita del Excel gana. Si no vino (o
        # vino vacía), se usa el decimal que se desprende del precio: así un
        # Excel cargado a mano con "1234,50" en una sola columna sigue
        # partiendo bien, sin obligar a llenar las siete columnas de
        # decimales a mano.
        decimal_explicito = normalize_decimal(celda(decimal_var)) if decimal_var in h else ""
        result[decimal_var] = decimal_explicito or decimal_auto

    # -- Texto: passthrough ----------------------------------------------
    for var in CANONICAL_VARS:
        if var in result:
            continue
        valor = celda(var)
        result[var] = "" if valor is None else str(valor).strip()

    # -- Campos internos (no se dibujan, deciden comportamiento) ----------
    internos = {}
    for campo in INTERNAL_FIELDS:
        valor = celda(campo)
        internos[campo] = "" if valor is None else str(valor).strip()

    # -- Vigencia: la fila gana, el campo global es el fallback -----------
    if not result["vigencia"]:
        result["vigencia"] = vigencia

    # -- Legales ----------------------------------------------------------
    # Solo se sustituyen si la persona los habilitó: muchas plantillas ya
    # traen el texto legal impreso en el diseño, y escribir encima duplica
    # la leyenda. Con el checkbox apagado la variable queda vacía y el
    # renderer no toca ese cuadro.
    if usar_legales:
        texto_legal = result["legales"] or legales
        if categoria_es_alcohol(internos["categoria"]):
            # Se SUMA, no pisa: la leyenda de alcohol es obligatoria y
            # convive con lo que haya escrito la persona.
            texto_legal = f"{texto_legal} {LEGAL_ALCOHOL}".strip() if texto_legal else LEGAL_ALCOHOL
        result["legales"] = texto_legal
    else:
        result["legales"] = ""

    return result


# ---------------------------------------------------------------------------
# Carga desde Excel
# ---------------------------------------------------------------------------

def load_products_from_bytes(
    excel_bytes: bytes,
    vigencia: str = "",
    legales: str = "",
    usar_legales: bool = False,
) -> list[dict]:
    """Lee el Excel de cenefas y devuelve una fila de variables por producto.

    Las columnas se matchean por nombre canónico normalizado y nada más: no
    hay alias. Una columna que no corresponda a ninguna variable se ignora
    en silencio (los exports traen decenas de columnas ajenas).
    """
    return _leer(excel_bytes, vigencia, legales, usar_legales)[0]


def load_products_con_headers(
    excel_bytes: bytes,
    vigencia: str = "",
    legales: str = "",
    usar_legales: bool = False,
) -> tuple[list[dict], list[str]]:
    """Igual que load_products_from_bytes, pero además devuelve los encabezados
    crudos del Excel.

    Los necesita la revisión previa: para avisar "la columna «Precio Anterior»
    no se reconoce" hay que saber qué columnas venían, no solo cuáles se
    entendieron.
    """
    return _leer(excel_bytes, vigencia, legales, usar_legales)


def _leer(
    excel_bytes: bytes, vigencia: str, legales: str, usar_legales: bool,
) -> tuple[list[dict], list[str]]:
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb["Cenefas"] if "Cenefas" in wb.sheetnames else wb.active

    # -- Detectar fila de encabezados -------------------------------------
    # El export real puede traer una fila de título y una en blanco antes de
    # los headers, así que no se asume la fila 1.
    header_row = None
    for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
        encontrados = {resolve(c) for c in row if c is not None}
        if encontrados & _DETECTION_VARS:
            header_row = i
            break
    if header_row is None:
        header_row = 1

    # -- Mapear columnas --------------------------------------------------
    h: dict[str, int] = {}
    for idx, cell in enumerate(ws[header_row]):
        canonica = resolve(cell.value) if cell.value is not None else None
        if canonica is None:
            continue
        # La primera columna con ese nombre gana. Caso real: el export de
        # Parrilla y Vinos trae "OFERTA" dos veces con significados
        # distintos, y la segunda pisaba a la primera en silencio.
        h.setdefault(canonica, idx)

    desc_col = h.get("descripcion")
    codigo_col = h.get("codigo")

    products: list[dict] = []
    seen: set = set()

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        # Una fila sin descripción no es un producto (filas de corte,
        # totales, o el relleno vacío que arrastra el export).
        if desc_col is not None and (desc_col >= len(row) or not row[desc_col]):
            continue
        if desc_col is None and codigo_col is not None and (codigo_col >= len(row) or not row[codigo_col]):
            continue

        data = process_row(row, h, vigencia=vigencia, legales=legales, usar_legales=usar_legales)

        # Deduplicación por código cuando existe -- es el identificador real
        # del producto. Sin código, se compara por la combinación de campos
        # visibles, único criterio disponible.
        codigo = (data.get("codigo") or "").strip()
        key = codigo or (
            (data.get("descripcion") or "").lower().strip(),
            data.get("precioOferta", ""),
            data.get("mecanica", ""),
        )
        if key in seen:
            continue
        seen.add(key)
        products.append(data)

    headers = [str(c.value).strip() if c.value is not None else ""
               for c in ws[header_row]]
    return products, headers


# ---------------------------------------------------------------------------
# Plantilla Excel de descarga
# ---------------------------------------------------------------------------
#
# Una sola plantilla para todos los destinos: el vocabulario de variables es
# el mismo en todos y cada diseño usa las que necesita. Antes había una
# plantilla por destino con columnas distintas, que es exactamente lo que
# esta unificación viene a eliminar.

_DOC = [
    ("codigo",               "Código de artículo / SKU",                       "Texto"),
    ("descripcion",          "Nombre del producto",                            "Texto"),
    ("mecanica",             "Mecánica de la oferta ya redactada",             "Texto"),
    ("tipoOferta",           "Titular de la oferta (ej. 2x1, 25% OFF)",        "Texto"),
    ("tipoOfertaComprando",  "Cuántas se llevan (ej. Comprando 2)",            "Texto"),
    ("unidad",               "La palabra que va debajo del precio (unidad)",   "Texto"),
    ("precioRegular",        "Precio regular / anterior — parte entera",       "Precio"),
    ("decimalPrecioRegular", "Decimales de precioRegular, con coma",           "Decimal"),
    ("precioOferta",         "Precio de oferta — parte entera",                "Precio"),
    ("decimalPrecioOferta",  "Decimales de precioOferta, con coma",            "Decimal"),
    ("promoOferta",          "Literal de la mecánica que tapa al precio (6x4)", "Precio"),
    ("decimalPromoOferta",   "Decimales de promoOferta, con coma",             "Decimal"),
    ("ofertaUno",            "Nivel de oferta 1 (ej. 3x) — número o texto",    "Precio"),
    ("decimalPrecioUno",     "Decimales de ofertaUno, con coma",               "Decimal"),
    ("ofertaDos",            "Nivel de oferta 2",                              "Precio"),
    ("decimalPrecioDos",     "Decimales de ofertaDos, con coma",               "Decimal"),
    ("ofertaTres",           "Nivel de oferta 3",                              "Precio"),
    ("decimalPrecioTres",    "Decimales de ofertaTres, con coma",              "Decimal"),
    ("ofertaCuatro",         "Nivel de oferta 4",                              "Precio"),
    ("decimalPrecioCuatro",  "Decimales de ofertaCuatro, con coma",            "Decimal"),
    ("precioBanco",          "Precio con beneficio bancario — parte entera",   "Precio"),
    ("decimalPrecioBanco",   "Decimales de precioBanco, con coma",             "Decimal"),
    ("banco",                "Nombre del banco o beneficio",                   "Texto"),
    ("vigencia",             "Período de validez de la promo",                 "Texto"),
    ("aclaracionUno",        "Primera aclaración",                             "Texto"),
    ("aclaracionDos",        "Segunda aclaración",                             "Texto"),
    ("aclaracionTres",       "Tercera aclaración",                             "Texto"),
    ("legales",              "Legales — solo se usan si se tildan al generar", "Texto"),
    ("dia",                  "Día",                                            "Texto"),
    ("mes",                  "Mes",                                            "Texto"),
    ("año",                  "Año",                                            "Texto"),
]

assert tuple(n for n, _, _ in _DOC) == ORDEN_EXPORT, (
    "_DOC quedó desincronizado de ORDEN_EXPORT: "
    f"faltan {set(ORDEN_EXPORT) - {n for n, _, _ in _DOC}}, "
    f"sobran {{n for n, _, _ in _DOC}} - set(ORDEN_EXPORT)"
)

_VIG = "Válido del 24.8 al 27.9"
_ACL = "Precio válido en todos los locales."

# Una fila por mecánica real, para que la plantilla se explique sola.
_EJEMPLOS = [
    #        codigo    descripcion                   mecanica                       tipoOferta  comprando      unidad    regular  dec  oferta  dec    ofertaUno  ...resto vacio...        vigencia  aclaracionUno
    ("610389", "ALFOMBRAS X4 GOMA AUTO",     "Precio Final",                "",       "",            "",       "960", "", "899", "", "", "",    "",   "", "", "", "", "", "", "", "", "", "", _VIG, _ACL, "", "", "", "", "", ""),
    ("599059", "COPA DE VIDRIO 500 ML",      "$74,50 la unidad.",           "2x1",    "",            "",       "149", "", "74", ",50", "2x1", "",    "",   "", "", "", "", "", "", "", "", "", "", _VIG, _ACL, "", "", "", "", "", ""),
    ("506232", "ATADO DE LEÑA 3KG",          "Comprando 3, $33 la unidad.", "",       "",            "",       "129", "", "99",  "", "", "",    "3x", "", "", "", "", "", "", "", "", "", "", _VIG, _ACL, "", "", "", "", "", ""),
    ("608093", "SET DE COLCHA Q 218X218 CM", "Precio Final",                "25% OFF","",            "",       "899", "", "719", ",20", "", "", "",   "", "", "", "", "", "", "", "", "", "", _VIG, _ACL, "", "", "", "", "", ""),
    # Rompe del Finde: la misma mecanica repartida en tres cuadros en vez de un
    # renglon. La cocarda lleva el literal, "Comprando N" va arriba del precio y
    # "unidad" abajo. Ahi `mecanica` queda vacia.
    ("172029", "CERVEZA STELLA ARTOIS 330ML","",                            "6x4",    "Comprando 6", "unidad", "113", "", "75",  ",33", "6x4", "", "",   "", "", "", "", "", "", "", "", "", "", _VIG, _ACL, "", "", "", "", "", ""),
]

# Cada ejemplo tiene que tener un valor por columna: si sobra o falta uno, la
# plantilla sale con los datos corridos de lugar y nadie lo nota hasta que la
# cenefa imprime el precio en el cuadro de la descripcion.
for _fila in _EJEMPLOS:
    assert len(_fila) == len(_DOC), (
        f"la fila de ejemplo {_fila[0]} tiene {len(_fila)} valores y hacen falta {len(_DOC)}"
    )



def generate_template_bytes(destino: str | None = None) -> bytes:
    """Plantilla Excel vacía con las 26 columnas del sistema.

    `destino` se acepta y se ignora — queda por compatibilidad con los
    callers existentes; ya no hay una plantilla distinta por destino.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    headers = [nombre for nombre, _, _ in _DOC]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cenefas"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    even_fill = PatternFill("solid", fgColor="EEF2F7")

    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, fila in enumerate(_EJEMPLOS, 2):
        for col_idx, value in enumerate(fila, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = even_fill

    for col, (nombre, _, _) in enumerate(_DOC, 1):
        ancho = 34 if nombre in ("descripcion", "mecanica", "vigencia", "legales") else 20
        ws.column_dimensions[get_column_letter(col)].width = ancho
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    # -- Hoja de referencia ------------------------------------------------
    ws2 = wb.create_sheet("Variables")
    for letra, ancho in (("A", 24), ("B", 52), ("C", 14)):
        ws2.column_dimensions[letra].width = ancho

    for col, name in enumerate(["Variable", "Qué es", "Tipo"], 1):
        cell = ws2.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 24

    for row_idx, fila in enumerate(_DOC, 2):
        for col_idx, value in enumerate(fila, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = even_fill

    nota = ws2.cell(
        row=len(_DOC) + 3, column=1,
        value="Ninguna columna es obligatoria. El nombre de la columna, el placeholder "
              "de la PPT (<<variable>>) y la variable son siempre el mismo texto.",
    )
    nota.font = Font(italic=True, size=10, color="475569")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
