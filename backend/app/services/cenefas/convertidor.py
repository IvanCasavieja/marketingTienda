"""Convertidor de Excel — matchea el export crudo de gestión contra el
catálogo compartido de descripciones (sku_descripciones) y arma un Excel
limpio, listo para subir directo al generador de Cenefas existente.

Deliberadamente separado de data_engine.py: esta herramienta no interpreta
la mecánica de oferta/oferta det (eso sigue siendo trabajo exclusivo del
generador de Cenefas cuando el Excel de salida se vuelva a subir ahí) —
solo transporta esas columnas tal cual del input al output.
"""
import csv
import difflib
import io
import re
import unicodedata
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.convertidor_header_alias import ConvertidorHeaderAlias
from app.models.sku_descripcion import SkuDescripcion
from app.services.cenefas.convertidor_ai import resolve_date_columns_with_ai
from app.services.cenefas.convertidor_variables import construir_variables
from app.services.cenefas.variables import ORDEN_EXPORT
from app.services.cenefas.formatters import parse_price_raw
from app.services.cenefas.validation_engine import DESCRIPTION_MAX_CHARS, DESCRIPTION_WARN_CHARS

# ---------------------------------------------------------------------------
# Normalización de headers del Excel de entrada
# ---------------------------------------------------------------------------

class ConvertidorParseError(ValueError):
    """El Excel no tiene una columna CODIGO reconocible."""


def _norm(name) -> str:
    """Copia intencional de data_engine._norm — no importamos ese símbolo
    privado entre módulos para no acoplar convertidor.py a los internos de
    data_engine.py; son 3 líneas, duplicarlas es más barato que el acople."""
    s = unicodedata.normalize("NFD", str(name)).encode("ascii", "ignore").decode()
    return re.sub(r"[\s_\-]+", "", s).lower()


_INPUT_ALIASES: dict[str, str] = {
    "codigo":         "codigo",
    "nombrearticulo": "nombre_articulo",
    # La columna "Descripción" cruda de gestión NO se mapea a propósito: no
    # pasó nunca por las reglas de estilo de la IA, así que no es una fuente
    # confiable para resolver ni para aprender en el catálogo compartido
    # (ver match_rows() -- un SKU sin match en el catálogo queda sin
    # descripción, para resolverse por "Generar con IA" o a mano).
    "moneda":         "moneda",
    "precioant":      "precio_anterior",
    "precio":         "precio",
    "oferta":         "oferta",
    "ofertadet":      "oferta_det",
    "descripcionweb": "descripcion_web",
    "comprador":      "comprador",
    "descuentoprov":     "descuento",
    "descuentoprovdet":  "descuento_det",
    # Fecha inicio/fin de vigencia -- ninguna de las dos es obligatoria en
    # gestión (muchos exports no las traen); si falta una o las dos, vigencia
    # simplemente queda como antes (ver _format_vigencia). Varios alias por
    # lado porque el nombre exacto de la columna varía según el export.
    "fechainicio":        "fecha_inicio",
    "fechadeinicio":      "fecha_inicio",
    "fechainicial":       "fecha_inicio",
    "fechadesde":         "fecha_inicio",
    "vigenciadesde":      "fecha_inicio",
    "iniciovigencia":     "fecha_inicio",
    "fechafin":           "fecha_fin",
    "fechadefin":         "fecha_fin",
    "fechafinal":         "fecha_fin",
    "fechahasta":         "fecha_fin",
    "vigenciahasta":      "fecha_fin",
    "finvigencia":        "fecha_fin",
}

_HEADER_SCAN_ROWS = 10
_DATE_SAMPLE_ROWS = 8  # filas de datos a mirar para juntar valores de muestra para la IA


def normalize_sku(raw) -> str:
    """int/float/str crudo de la celda CODIGO -> string canónico ('17780.0' -> '17780')."""
    if raw is None:
        return ""
    if isinstance(raw, float):
        raw = int(raw) if raw.is_integer() else raw
    s = str(raw).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


async def upsert_sku_descripcion(db: AsyncSession, sku: str, descripcion: str, user_id: int) -> str:
    """Upsert vía ON CONFLICT DO UPDATE — evita una condición de carrera real
    si dos personas completan el mismo SKU sin match al mismo tiempo.
    Compartido entre el PATCH manual del Convertidor y la tool de Tinín, para
    no duplicar el statement en dos lugares. No commitea — el caller decide
    cuándo (el PATCH lo hace solo, Tinín puede encadenar varias llamadas en
    un mismo turno de tool-use antes de commitear una vez)."""
    sku_norm = normalize_sku(sku)
    if not sku_norm:
        raise ValueError("SKU inválido")
    descripcion = descripcion.strip()
    if not descripcion:
        raise ValueError("La descripción no puede quedar vacía")
    descripcion = descripcion[:300]

    stmt = pg_insert(SkuDescripcion).values(
        sku=sku_norm,
        descripcion=descripcion,
        updated_by_id=user_id,
    ).on_conflict_do_update(
        index_elements=["sku"],
        set_={
            "descripcion": descripcion,
            "updated_by_id": user_id,
            "updated_at": func.now(),
        },
    )
    await db.execute(stmt)
    return sku_norm


def _parse_price_or_none(raw) -> float | None:
    if raw is None or str(raw).strip() == "":
        return None
    # parse_price_raw exige que el string empiece con un dígito (ver
    # formatters.py) -- un precio tipo "$150,50" no matchea y devolvería
    # 0.0 en silencio. _is_numeric_like sí tolera un "$" adelante, así que
    # sin este strip la validación diría "está bien" mientras el valor
    # real usado en la fila/export quedaba en cero. Mismo patrón que ya
    # usa data_engine.py antes de llamar a esta misma función.
    clean = re.sub(r"^[^\d]*", "", str(raw).strip())
    return parse_price_raw(clean) if clean else 0.0


def _clean_str(raw) -> str:
    return str(raw).strip() if raw is not None else ""


_DATE_FORMATS = ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y")


def _parse_date_or_none(raw) -> date | None:
    """xlsx con la celda formateada como fecha llega ya como date/datetime
    (openpyxl con data_only=True); CSV y celdas de texto llegan como string
    en alguno de los formatos regionales más comunes. Cualquier otra cosa
    (vacío, texto que no matchea ningún formato) es None -- fecha_inicio/fin
    son opcionales, así que un valor no reconocible se ignora en vez de
    romper el import entero."""
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip() if raw is not None else ""
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


_MESES = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]


def _format_vigencia(fecha_inicio: date | None, fecha_fin: date | None) -> str:
    """"Desde el 10 al 16 de julio" (mismo mes/año -- el caso común de una
    promo semanal), con fallback a mencionar el mes de cada punta si difieren.
    Ninguna de las dos fechas es obligatoria: con una sola, frase abierta
    ("Desde"/"Hasta" nomás); con ninguna, "" (igual que antes de esta regla,
    completable a mano en la grilla)."""
    if fecha_inicio and fecha_fin:
        mes_inicio = _MESES[fecha_inicio.month - 1]
        if fecha_inicio.year != fecha_fin.year:
            mes_fin = _MESES[fecha_fin.month - 1]
            return (
                f"Desde el {fecha_inicio.day} de {mes_inicio} de {fecha_inicio.year} "
                f"hasta el {fecha_fin.day} de {mes_fin} de {fecha_fin.year}"
            )
        if fecha_inicio.month != fecha_fin.month:
            mes_fin = _MESES[fecha_fin.month - 1]
            return f"Desde el {fecha_inicio.day} de {mes_inicio} hasta el {fecha_fin.day} de {mes_fin}"
        return f"Desde el {fecha_inicio.day} al {fecha_fin.day} de {mes_inicio}"
    if fecha_inicio:
        return f"Desde el {fecha_inicio.day} de {_MESES[fecha_inicio.month - 1]}"
    if fecha_fin:
        return f"Hasta el {fecha_fin.day} de {_MESES[fecha_fin.month - 1]}"
    return ""


# ---------------------------------------------------------------------------
# Validación por tipo esperado de columna
# ---------------------------------------------------------------------------
#
# Cada columna del Excel de gestión tiene un tipo de contenido esperado —
# precio es numérico, nombre/descripción son texto, moneda es un símbolo de
# un set chico conocido, oferta det es una categoría (nunca un número). Si
# el valor de una celda no matchea el tipo esperado de SU columna, eso es la
# señal real de columnas corridas para esa fila — no una heurística de
# "pinta" sobre la fila entera, sino una razón concreta y localizada: "acá
# esperaba X y encontré Y". OFERTA queda sin validar a propósito: en los
# datos reales es legítimamente polimórfica (texto "PVP OFERTA", un precio
# repetido, o una mecánica "2x599"), no hay un tipo único que reclamarle.

_VALID_MONEDAS = {"$", "u$s", "us$", "usd", "uyu"}
_NUMERIC_RE  = re.compile(r"^[\$\s]*-?\d[\d.,]*\s*$")
_LETTER_RE   = re.compile(r"[^\W\d_]")  # al menos una letra (unicode-aware)


def _is_numeric_like(raw: str) -> bool:
    return bool(_NUMERIC_RE.match(raw.strip()))


def _has_letters(raw: str) -> bool:
    return bool(_LETTER_RE.search(raw))


# ---------------------------------------------------------------------------
# Parseo del Excel de entrada
# ---------------------------------------------------------------------------

def _read_csv_rows(csv_bytes: bytes) -> list[tuple]:
    """Lee un CSV (algunos exports de gestión, sobre todo Rompe Precios,
    vienen así en vez de xlsx) tolerando las dos variantes regionales más
    comunes: separador coma o punto y coma (frecuente acá, porque la coma
    ya se usa como separador decimal), y encoding UTF-8 o Windows-1252
    (típico de sistemas de gestión viejos que no exportan UTF-8 nativo)."""
    text = None
    for encoding in ("utf-8-sig", "cp1252"):
        try:
            text = csv_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = csv_bytes.decode("utf-8", errors="replace")

    sample = text[:4096]
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter
    except csv.Error:
        # El sniffer necesita al menos un par de filas consistentes para
        # decidir — con muy pocas filas o un formato ambiguo, cae acá:
        # cuenta cuál separador aparece más seguido en la muestra.
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","

    return [tuple(row) for row in csv.reader(io.StringIO(text), delimiter=delimiter)]


def _parece_csv_en_una_columna(rows: list[tuple]) -> bool:
    """True si el .xlsx en realidad es un CSV metido en una sola columna.

    El export real de gestión sale así: una única columna por fila, con todo
    el contenido separado por comas adentro de la celda. openpyxl lo lee tal
    cual, así que el header queda como "CODIGO,SECCION,NRO_OFERTA,..." y no
    matchea con ninguna columna esperada -- el archivo se rechazaba con "no
    encontré una columna CODIGO" sin ninguna pista de por qué.
    """
    if not rows:
        return False
    con_datos = [r for r in rows[:5] if r and r[0] is not None]
    if not con_datos:
        return False
    if any(len(r) > 1 and any(c is not None for c in r[1:]) for r in con_datos):
        return False
    return all(str(r[0]).count(",") >= 3 or str(r[0]).count(";") >= 3 for r in con_datos)


def _mapear_columnas(row) -> tuple[dict[int, str], bool]:
    candidate: dict[int, str] = {}
    found_codigo = False
    for col_idx, cell in enumerate(row):
        if cell is None:
            continue
        var_name = _INPUT_ALIASES.get(_norm(cell))
        if var_name:
            candidate[col_idx] = var_name
            if var_name == "codigo":
                found_codigo = True
    return candidate, found_codigo


def detectar_fila_headers(rows: list[tuple]) -> int | None:
    """Indice (0-based) de la fila de encabezados, o None si no la hay.

    No se asume la fila 1: el export real trae una fila de titulo y una en
    blanco antes. La senal es una columna CODIGO reconocible."""
    for i, row in enumerate(rows[:_HEADER_SCAN_ROWS]):
        _, found_codigo = _mapear_columnas(row)
        if found_codigo:
            return i
    return None


def leer_filas(file_bytes: bytes, filename: str = "") -> list[tuple]:
    """Lee el archivo subido a filas, sea .csv, .xlsx o .xlsx-que-es-un-CSV."""
    if filename.lower().endswith(".csv"):
        return _read_csv_rows(file_bytes)

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=1, max_row=None, values_only=True))

    if _parece_csv_en_una_columna(rows):
        crudo = chr(10).join(str(r[0]) for r in rows if r and r[0] is not None)
        return _read_csv_rows(crudo.encode("utf-8"))
    return rows


async def parse_input_excel(
    file_bytes: bytes,
    filename: str = "",
    *,
    db: AsyncSession,
    current_user_id: int,
    allow_ai: bool = False,
    mapeo: dict[str, str] | None = None,
    valores: dict[str, str] | None = None,
) -> tuple[list[dict], int, list[str]]:
    """Detecta la fila de headers real (puede no ser la fila 1 — el export
    real de gestión trae una fila de título + una fila en blanco antes),
    mapea columnas por nombre normalizado, y extrae por fila: codigo,
    nombre_articulo, moneda, precio_anterior, precio, oferta, oferta_det,
    descripcion_web, comprador, descuento, descuento_det, fecha_inicio,
    fecha_fin (estas últimas dos opcionales -- ver _format_vigencia).

    Columnas no reconocidas por _INPUT_ALIASES se resuelven en dos pasos más
    antes de darse por ignoradas, solo para fecha_inicio/fecha_fin: primero
    contra ConvertidorHeaderAlias (headers que Tinín ya clasificó en un import
    anterior — nunca vuelve a gastar una llamada a IA en el mismo nombre de
    columna dos veces), y si sigue sin match y allow_ai=True (el caller decide
    esto según el permiso ai.tinin del usuario), se le pide a Tinín que
    clasifique las columnas no reconocidas cuyos valores de muestra ya
    parsean como fecha real (ver resolve_date_columns_with_ai en
    convertidor_ai.py) — nunca se le pregunta por columnas que no tienen
    pinta de fecha en los datos.

    Acepta .xlsx/.xlsm, .csv, y también el .xlsx que en realidad es un CSV
    metido en una sola columna, que es como sale el export real de gestión
    (ver leer_filas).

    Las variables que el Convertidor no puede deducir solo (ofertaUno..Cuatro,
    vigencia, aclaracionUno..Tres, legales) las resuelve la persona en la
    pantalla de mapeo, de una de dos formas:

    - `mapeo`   {variable: nombre_de_columna} -- se lee de esa columna, fila
      por fila.
    - `valores` {variable: texto_fijo} -- el mismo texto para todas las filas.
      Hace falta porque el export de gestión no trae nunca vigencia ni
      legales: esos textos los escribe una persona, no salen de ninguna
      columna.

    Las dos son excluyentes por variable; si igual llegaran las dos, gana el
    valor fijo (es lo que se escribió explícitamente para esta corrida).
    Ambas viajan resueltas en cada fila bajo la clave "_mapeado".

    Devuelve (filas, learned_aliases_count, headers) -- learned_aliases_count
    es cuántos headers nuevos aprendió Tinín en esta llamada, para que el
    caller sepa si hace falta commitear; headers son los nombres crudos de la
    fila de encabezados, para poder re-abrir la pantalla de mapeo."""
    rows = leer_filas(file_bytes, filename)

    header_row_idx = detectar_fila_headers(rows)
    col_map: dict[int, str] = {}
    if header_row_idx is not None:
        col_map, _ = _mapear_columnas(rows[header_row_idx])

    if header_row_idx is None:
        raise ConvertidorParseError(
            "No encontré una columna 'CODIGO' reconocible en las primeras "
            f"{_HEADER_SCAN_ROWS} filas del Excel — verificá que sea el export crudo de gestión."
        )

    learned_aliases_count = 0
    header_row = rows[header_row_idx]
    mapped_cols = set(col_map.keys())
    # col_idx de cada celda con texto en la fila de headers que _INPUT_ALIASES
    # no supo mapear -- candidatas a resolverse por el cache aprendido o por IA.
    unresolved_by_norm: dict[str, int] = {}
    unresolved_display: dict[str, str] = {}
    for col_idx, cell_val in enumerate(header_row):
        if cell_val is None or col_idx in mapped_cols:
            continue
        norm = _norm(cell_val)
        if not norm:
            continue
        unresolved_by_norm[norm] = col_idx
        unresolved_display[norm] = str(cell_val).strip()

    if unresolved_by_norm:
        result = await db.execute(
            select(ConvertidorHeaderAlias.header_norm, ConvertidorHeaderAlias.field_name)
            .where(ConvertidorHeaderAlias.header_norm.in_(unresolved_by_norm.keys()))
        )
        for header_norm, field_name in result.all():
            # Se saca de unresolved pase lo que pase -- un cache negativo
            # (field_name None, "confirmado que no es de vigencia") también
            # evita volver a preguntarle a la IA por este mismo header.
            col_idx = unresolved_by_norm.pop(header_norm, None)
            if col_idx is not None and field_name is not None:
                col_map[col_idx] = field_name

    if unresolved_by_norm and allow_ai:
        sample_rows = rows[header_row_idx + 1: header_row_idx + 1 + _DATE_SAMPLE_ROWS]
        candidates = []
        for header_norm, col_idx in unresolved_by_norm.items():
            muestras = [
                str(r[col_idx]).strip()
                for r in sample_rows
                if col_idx < len(r) and r[col_idx] is not None and _parse_date_or_none(r[col_idx])
            ]
            # Al menos dos valores de muestra parseando como fecha real -- uno
            # solo puede ser casualidad (un texto que por azar matchea un
            # formato de fecha), dos ya es señal sólida de que la columna
            # entera es de fechas y vale la pena preguntarle a Tinín cuál es.
            if len(muestras) >= 2:
                candidates.append({
                    "header_norm": header_norm,
                    "header_display": unresolved_display[header_norm],
                    "muestras": muestras[:3],
                })

        if candidates:
            # {} significa que la llamada entera falló (red, JSON con forma
            # rara, etc.) -- no cachear nada en ese caso, se reintenta en el
            # próximo import. Un dict no vacío trae una entrada por candidato
            # (positiva o None), ver resolve_date_columns_with_ai.
            clasificaciones = await resolve_date_columns_with_ai(candidates, db, current_user_id)
            if clasificaciones:
                nuevas_aliases: list[tuple[str, str | None]] = []
                seen_fields: set[str] = set()
                for header_norm, field_name in clasificaciones.items():
                    if field_name is not None:
                        if field_name in seen_fields:
                            # Dos columnas distintas no pueden ser el mismo
                            # campo -- ante la duda no asignamos ninguna de
                            # las dos (mejor perder la detección que pisar
                            # una con la otra), y se cachea como "no es de
                            # vigencia" para no reabrir la duda en el
                            # próximo import con el mismo layout.
                            field_name = None
                        else:
                            seen_fields.add(field_name)
                            col_idx = unresolved_by_norm.get(header_norm)
                            if col_idx is not None:
                                col_map[col_idx] = field_name
                    nuevas_aliases.append((header_norm, field_name))

                stmt = pg_insert(ConvertidorHeaderAlias).values([
                    {"header_norm": h, "field_name": f} for h, f in nuevas_aliases
                ]).on_conflict_do_nothing(index_elements=["header_norm"])
                await db.execute(stmt)
                learned_aliases_count = len(nuevas_aliases)

    # col_map es col_idx -> var_name; invertido una sola vez, no por fila.
    col_by_var = {var: c for c, var in col_map.items()}
    codigo_col = col_by_var["codigo"]

    def cell(row: tuple, var: str):
        c = col_by_var.get(var)
        return row[c] if c is not None and c < len(row) else None

    # {variable: col_idx} para lo que eligió la persona en la pantalla de
    # mapeo. Se resuelve por nombre normalizado para que un espacio o una
    # mayúscula de más no rompa el match contra el header real.
    headers_crudos = [str(c).strip() if c is not None else "" for c in header_row]
    por_norm = {_norm(h): i for i, h in enumerate(headers_crudos) if h}
    mapeo_cols: dict[str, int] = {}
    for var, col_nombre in (mapeo or {}).items():
        idx = por_norm.get(_norm(col_nombre or ""))
        if idx is not None:
            mapeo_cols[var] = idx

    # Valores escritos a mano: el mismo texto en todas las filas. Se limpian
    # acá una sola vez en vez de por fila.
    fijos = {
        var: str(val).strip()
        for var, val in (valores or {}).items()
        if val is not None and str(val).strip()
    }

    parsed: list[dict] = []
    for row in rows[header_row_idx + 1:]:
        # "not row[...]" en vez de "is None": una celda vacía de CSV llega
        # como "" (nunca None, a diferencia de openpyxl) — este chequeo
        # cubre las dos fuentes por igual.
        if codigo_col >= len(row) or not row[codigo_col]:
            continue
        codigo = normalize_sku(row[codigo_col])
        if not codigo:
            continue

        precio_raw          = _clean_str(cell(row, "precio"))
        precio_anterior_raw = _clean_str(cell(row, "precio_anterior"))
        parsed.append({
            "codigo":            codigo,
            "nombre_articulo":   _clean_str(cell(row, "nombre_articulo")),
            "moneda":            _clean_str(cell(row, "moneda")) or "$",
            "precio_anterior":   _parse_price_or_none(precio_anterior_raw),
            "precio_anterior_raw": precio_anterior_raw,
            "precio":            _parse_price_or_none(precio_raw),
            "precio_raw":        precio_raw,
            "oferta":            _clean_str(cell(row, "oferta")),
            "oferta_det":        _clean_str(cell(row, "oferta_det")),
            "descripcion_web":   _clean_str(cell(row, "descripcion_web")),
            "comprador":         _clean_str(cell(row, "comprador")),
            "descuento":         _clean_str(cell(row, "descuento")),
            "descuento_det":     _clean_str(cell(row, "descuento_det")),
            "fecha_inicio":      _parse_date_or_none(cell(row, "fecha_inicio")),
            "fecha_fin":         _parse_date_or_none(cell(row, "fecha_fin")),
            "_mapeado": {
                **{var: _clean_str(row[i]) if i < len(row) else ""
                   for var, i in mapeo_cols.items()},
                **fijos,
            },
        })
    return parsed, learned_aliases_count, headers_crudos


# ---------------------------------------------------------------------------
# Fiambres por kg -> deben ir a 100g (descripción) y precio÷10
# ---------------------------------------------------------------------------
#
# Mismo criterio que ya usan dos reglas independientes del generador de
# Cenefas (data_engine.py): "fiambr" en COMPRADOR es la señal de categoría
# (ese precedente vive en process_row, líneas ~222-236), y "kg" como unidad
# suelta en el texto es la señal de que todavía está en kilogramo (ese
# criterio viene del precedente de subCategoria en _apply_legacy_compute).
# Acá se combinan: a diferencia de ambos precedentes, esto NO calcula el
# precio nuevo ni lo persiste en ningún lado — solo marca la fila para que
# el frontend decida qué mostrar/sugerir.
#
# Excepción confirmada con el equipo: chorizo, frankfurters y panchos NUNCA
# pasan a 100g aunque el texto diga "Kg" -- el chorizo se vende por kilo por
# decisión comercial, y frankfurters/panchos mantienen el peso original del
# envase. Ninguno de los dos casos es "menos fiambre" que jamón cocido o
# salame (que sí convierten) -- es una excepción por línea de producto, no
# por categoría.

_RE_FIAMBRE = re.compile(r"fiambr", re.IGNORECASE)
_RE_UNIDAD_KG = re.compile(r"(?:^|[\s.])kg\.?(?:$|[\s.,)])", re.IGNORECASE)
_RE_SIN_CONVERSION_100G = re.compile(r"\bchorizos?\b|\bfrankfurters?\b|\bpanchos?\b", re.IGNORECASE)


def _tiene_unidad_kg(*textos: str) -> bool:
    return any(_RE_UNIDAD_KG.search(t) for t in textos if t)


def _tiene_producto_sin_conversion(*textos: str) -> bool:
    return any(_RE_SIN_CONVERSION_100G.search(t) for t in textos if t)


def _es_fiambre_por_kg(comprador: str, nombre_articulo: str, descripcion: str, descripcion_web: str) -> bool:
    if not comprador or not _RE_FIAMBRE.search(comprador):
        return False
    if _tiene_producto_sin_conversion(nombre_articulo, descripcion, descripcion_web):
        return False
    return _tiene_unidad_kg(nombre_articulo, descripcion, descripcion_web)


# ---------------------------------------------------------------------------
# Matching contra el catálogo + warnings
# ---------------------------------------------------------------------------

def _compute_warnings(row: dict) -> list[str]:
    """Un warning por columna, cada uno con su propio motivo.

    Dos familias: "falta el dato" (missing_*) y "hay contenido pero no del
    tipo que esa columna espera" (*_invalido), que es la señal real de un
    Excel con las columnas corridas, localizada en la columna exacta que no
    cierra. Se suman los warnings de mecánica que ya calculó
    convertidor_variables (oferta_inesperada, combo_no_parseable...), que
    viajan en la fila bajo "warnings_mecanica".
    """
    w = list(row.get("warnings_mecanica") or [])

    descripcion = str(row.get("descripcion") or "").strip()
    if not descripcion:
        w.append("missing_description")
    elif not _has_letters(descripcion):
        w.append("descripcion_invalida")
    elif len(descripcion) > DESCRIPTION_MAX_CHARS:
        # Es para un cartel de precio, no un párrafo -- mismos umbrales que
        # validation_engine.py, así no hay que inventar un límite nuevo acá.
        # Ahora pesa más que antes: el motor ya no achica la descripción sola
        # para que entre (ver la nota en component_renderer.py).
        w.append("descripcion_larga")
    elif len(descripcion) > DESCRIPTION_WARN_CHARS:
        w.append("descripcion_algo_larga")

    precio_raw = str(row.get("precio_raw") or "").strip()
    if not precio_raw:
        w.append("missing_price")
    elif not _is_numeric_like(precio_raw):
        w.append("precio_invalido")

    precio_anterior_raw = str(row.get("precio_anterior_raw") or "").strip()
    if not precio_anterior_raw:
        w.append("missing_precio_anterior")
    elif not _is_numeric_like(precio_anterior_raw):
        w.append("precio_anterior_invalido")

    oferta_det = str(row.get("oferta_det") or "").strip()
    if not oferta_det:
        w.append("missing_oferta_det")
    elif _is_numeric_like(oferta_det):
        w.append("oferta_det_invalido")  # es una categoría, nunca un número puro

    descripcion_web = str(row.get("descripcion_web") or "").strip()
    if not descripcion_web:
        w.append("missing_descripcion_web")
    elif not _has_letters(descripcion_web):
        w.append("descripcion_web_invalida")

    moneda = str(row.get("moneda") or "").strip()
    moneda_norm = moneda.lower()
    if moneda_norm and moneda_norm not in _VALID_MONEDAS:
        w.append("moneda_invalida")
    elif moneda_norm and moneda_norm not in ("$", "uyu"):
        # Desde 08/2026 el símbolo de moneda es texto FIJO del diseño de la
        # PPT (dejó de ser una variable), así que una fila en dólares se
        # imprimiría con el "$" de la plantilla y un precio que no es en
        # pesos. No se puede arreglar solo: lo tiene que ver una persona.
        w.append("moneda_no_pesos")

    nombre_articulo = str(row.get("nombre_articulo") or "").strip()
    if nombre_articulo and not _has_letters(nombre_articulo):
        w.append("nombre_articulo_invalido")

    return w


def _chunks(items: list, size: int):
    for i in range(0, len(items), size):
        yield items[i:i + size]


# ---------------------------------------------------------------------------
# Pares "mismo producto, dos SKUs" -- patrón real: sufijo suelto " M" / " A"
# al final del nombre (ej. "CUADRIL S/HUESO M" / "CUADRIL SIN HUESO A"), no
# exclusivo de una categoría -- confirmado también en NALGA con datos reales.
# ---------------------------------------------------------------------------

_MA_SIMILARITY_THRESHOLD = 0.6
# Sufijo M/A suelto al final del nombre, con o sin comillas -- "COCA COLA" NO
# matchea (la "A" es parte de la palabra, no un sufijo separado por espacio).
_RE_MA_SUFFIX = re.compile(r'(?:^|\s)["\']*([MA])["\']*\s*$', re.IGNORECASE)


def _ma_base_and_suffix(nombre_articulo: str) -> tuple[str, str] | None:
    s = nombre_articulo.strip()
    m = _RE_MA_SUFFIX.search(s)
    if not m:
        return None
    base = s[:m.start()].strip()
    return (base, m.group(1).upper()) if base else None


def detect_ma_pairs(rows: list[dict], catalogo: dict[str, str]) -> list[dict]:
    """Agrupa candidatos (nombre terminado en sufijo suelto M/A) por
    comprador, y empareja greedy por similarity ratio (difflib) sobre la base
    sin sufijo -- no exige igualdad exacta: "S/HUESO" vs "SIN HUESO" ~0.82.
    Un par necesita un M y un A (nunca dos del mismo sufijo) del mismo
    comprador. Saltea pares donde ambos SKUs ya resuelven a la misma
    descripción en el catálogo (ya unificados antes, incluyendo por la
    segunda pasada de fallback de SKU compuesto en match_rows)."""
    candidatos: dict[str, list[dict]] = {}
    for r in rows:
        parsed = _ma_base_and_suffix(r["nombre_articulo"])
        if not parsed:
            continue
        base, suffix = parsed
        comprador = (r.get("comprador") or "").strip().lower()
        candidatos.setdefault(comprador, []).append({
            "codigo": r["codigo"],
            "nombre_articulo": r["nombre_articulo"],
            "base": base,
            "suffix": suffix,
        })

    pairs: list[dict] = []
    seen_skus: set[str] = set()
    for items in candidatos.values():
        used: set[int] = set()
        for i, a in enumerate(items):
            if i in used or a["codigo"] in seen_skus:
                continue
            best_j, best_ratio = None, 0.0
            for j in range(i + 1, len(items)):
                b = items[j]
                if j in used or b["suffix"] == a["suffix"] or b["codigo"] in seen_skus:
                    continue
                ratio = difflib.SequenceMatcher(None, a["base"].lower(), b["base"].lower()).ratio()
                if ratio > best_ratio:
                    best_ratio, best_j = ratio, j
            if best_j is None or best_ratio < _MA_SIMILARITY_THRESHOLD:
                continue
            b = items[best_j]
            if catalogo.get(a["codigo"]) and catalogo.get(a["codigo"]) == catalogo.get(b["codigo"]):
                continue  # ya unificados antes -- mismo SKU compuesto ya resuelve a ambos
            used.add(i)
            used.add(best_j)
            seen_skus.add(a["codigo"])
            seen_skus.add(b["codigo"])
            pairs.append({
                "sku1": a["codigo"],
                "sku2": b["codigo"],
                "nombre1": a["nombre_articulo"],
                "nombre2": b["nombre_articulo"],
                "base": a["base"] if len(a["base"]) >= len(b["base"]) else b["base"],
            })
    return pairs


async def match_rows(
    parsed: list[dict], db: AsyncSession
) -> tuple[list[dict], list[dict]]:
    """Bulk lookup por SKU (un SELECT por cada 1000 códigos distintos, no
    N queries) + cómputo de warnings por fila.

    La descripción de una fila sale ÚNICA Y EXCLUSIVAMENTE del catálogo
    (sku_descripciones) -- nunca de la columna "Descripción" cruda que trae
    el Excel de gestión, aunque tenga contenido: esa columna no pasó nunca
    por las reglas de estilo de la IA (ver _STYLE_RULES en convertidor_ai.py)
    ni por revisión humana, así que no es una fuente confiable para el
    catálogo compartido. Un SKU sin match queda con descripción vacía
    (warning "missing_description", fila roja) para resolverse por
    "Generar con IA" (usa nombre_articulo + descripcion_web, sí curadas) o a
    mano -- nunca se aprende nada acá de forma automática/silenciosa.

    Devuelve (rows, ma_pairs) -- ma_pairs son los pares "mismo producto, dos
    SKUs" detectados (ver detect_ma_pairs) todavía sin unificar."""
    skus = sorted({r["codigo"] for r in parsed if r["codigo"]})
    catalogo: dict[str, str] = {}
    for chunk in _chunks(skus, 1000):
        result = await db.execute(
            select(SkuDescripcion.sku, SkuDescripcion.descripcion)
            .where(SkuDescripcion.sku.in_(chunk))
        )
        catalogo.update(dict(result.all()))

    # Fallback para SKUs sueltos que ya fueron unificados antes: si un par
    # M/A se unificó en una carga anterior, sku_descripciones solo tiene la
    # entrada compuesta "SKU1-SKU2" -- sin este paso, una carga futura con
    # cualquiera de los dos SKUs sueltos no matchearía más. Solo se busca lo
    # que hace falta (faltantes), no todos los SKUs de este import.
    faltantes = {s for s in skus if s not in catalogo}
    if faltantes:
        result = await db.execute(
            select(SkuDescripcion.sku, SkuDescripcion.descripcion).where(SkuDescripcion.sku.contains("-"))
        )
        for compuesto, desc in result.all():
            # .strip() por parte -- el merge M/A junta sin espacios ("SKU1-SKU2"),
            # pero "Unificar categorías" junta con " - " (espacio-guion-espacio,
            # ver ConvertidorGrid.tsx: commitUnificacion) para que se lea mejor con
            # varios SKUs en una celda. Sin el strip, la mitad de cada parte de esos
            # compuestos quedaría con un espacio colgado y nunca matchearía contra
            # faltantes (que son SKUs ya normalizados, sin espacios).
            for parte in compuesto.split("-"):
                parte = parte.strip()
                if parte in faltantes:
                    catalogo.setdefault(parte, desc)

    rows = []
    for i, r in enumerate(parsed):
        descripcion = catalogo.get(r["codigo"], "")

        # Las 26 variables ya resueltas: mecánica redactada, precios partidos
        # en entero + decimal, y lo que la persona mapeó pisando lo calculado.
        variables, warn_mecanica = construir_variables(
            r,
            descripcion,
            r.get("_mapeado") or {},
            vigencia_fallback=_format_vigencia(r.get("fecha_inicio"), r.get("fecha_fin")),
        )

        # Contexto del export de gestión: no son variables y no se exportan,
        # pero se muestran en la grilla para que una persona pueda entender
        # de dónde salió cada valor calculado y corregirlo si algo no cierra.
        contexto = {
            "nombre_articulo":     r["nombre_articulo"],
            "comprador":           r["comprador"],
            "moneda":              r["moneda"],
            "oferta_origen":       r["oferta"],
            "oferta_det":          r["oferta_det"],
            "descripcion_web":     r["descripcion_web"],
            "precio_raw":          r["precio_raw"],
            "precio_anterior_raw": r["precio_anterior_raw"],
        }

        fila = {
            "row_id":  i,
            "matched": bool(descripcion),
            **contexto,
            **variables,
            "es_fiambre_kg": _es_fiambre_por_kg(
                r["comprador"], r["nombre_articulo"], descripcion, r["descripcion_web"]
            ),
            "warnings_mecanica": warn_mecanica,
        }
        fila["warnings"] = _compute_warnings(fila)
        rows.append(fila)

    ma_pairs = detect_ma_pairs(rows, catalogo)

    return rows, ma_pairs


# ---------------------------------------------------------------------------
# Generación del Excel de salida
# ---------------------------------------------------------------------------

# El Excel de salida tiene UNA COLUMNA POR VARIABLE, con el nombre exacto de
# la variable como encabezado. No hay traducción ni nombres "bonitos": ese
# archivo se vuelve a subir al generador de cenefas, que matchea las columnas
# por nombre canónico y nada más. Antes la salida repetía las columnas de
# gestión (Oferta, Oferta Det, Descripción Web...) y el generador tenía que
# volver a interpretarlas; ahora los valores ya vienen resueltos.
_ANCHAS = {"descripcion", "mecanica", "vigencia", "legales",
           "aclaracionUno", "aclaracionDos", "aclaracionTres"}

# Columnas que salen SIEMPRE, tengan dato o no: son las que arman una cenefa
# mínima. Verlas vacías es justamente la señal de que falta completarlas, y
# los dos decimales acompañan a su precio porque el diseño de la cenefa tiene
# un cuadro aparte para ellos.
_COLUMNAS_FIJAS: frozenset[str] = frozenset((
    "codigo", "descripcion", "mecanica",
    "precioRegular", "decimalPrecioRegular",
    "precioOferta",  "decimalPrecioOferta",
))


def _columnas_de_salida(rows: list[dict]) -> list[str]:
    """Qué columnas lleva el Excel: las fijas más las que traen algún dato.

    Una columna que queda vacía en TODAS las filas no se crea. Antes salían
    las 26 siempre y un listado normal de gestión --que no trae vigencia, ni
    aclaraciones, ni niveles de oferta, ni banco-- se bajaba con 19 columnas
    vacías al lado de las que importan.

    No se pierde nada: el generador de cenefas no exige ninguna variable, y
    una que no está en el Excel simplemente no se sustituye en el diseño.
    """
    con_dato = {
        var for var in ORDEN_EXPORT
        if any(str(r.get(var, "") or "").strip() for r in rows)
    }
    return [v for v in ORDEN_EXPORT if v in con_dato or v in _COLUMNAS_FIJAS]


# Warning -> variable cuya columna se resalta. Se guarda el NOMBRE y no el
# indice porque el Excel de salida ya no lleva siempre las mismas columnas
# (ver _columnas_de_salida): el indice se resuelve recien al armarlo.
#
# Los de mecanica apuntan a la columna mecanica, que es donde se ve el
# resultado de la interpretacion que hay que revisar.
_WARN_VAR = {
    "missing_description":      "descripcion",
    "descripcion_invalida":     "descripcion",
    "descripcion_larga":        "descripcion",
    "descripcion_algo_larga":   "descripcion",
    "missing_price":            "precioOferta",
    "precio_invalido":          "precioOferta",
    "missing_precio_anterior":  "precioRegular",
    "precio_anterior_invalido": "precioRegular",
    "oferta_inesperada":        "mecanica",
    "combo_no_parseable":       "mecanica",
    "mxn_no_parseable":         "mecanica",
    "mxn_sin_precio":           "mecanica",
    "oferta_det_invalido":      "mecanica",
    "missing_oferta_det":       "mecanica",
    "moneda_invalida":          "precioOferta",
    "moneda_no_pesos":          "precioOferta",
}

# Warnings que NO son "falta el dato" sino "hay contenido que no cierra":
# apuntan a un Excel con las columnas corridas o a un valor que una persona
# tiene que decidir. Se pintan distinto porque no se arreglan completando.
_INVALID_TYPE_CODES = {
    "nombre_articulo_invalido", "descripcion_invalida", "descripcion_larga",
    "moneda_invalida", "moneda_no_pesos",
    "precio_anterior_invalido", "precio_invalido", "oferta_det_invalido",
    "descripcion_web_invalida",
    "oferta_inesperada", "combo_no_parseable", "mxn_no_parseable", "mxn_sin_precio",
}


def build_output_workbook(rows: list[dict]) -> bytes:
    fields = _columnas_de_salida(rows)
    headers = fields
    col_widths = [34 if v in _ANCHAS else 18 for v in fields]
    # {codigo_de_warning: indice_de_columna} solo para las columnas que
    # realmente salieron.
    warn_col = {
        code: fields.index(var) + 1
        for code, var in _WARN_VAR.items() if var in fields
    }
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cenefas"

    header_fill    = PatternFill("solid", fgColor="1E3A5F")
    header_font    = Font(bold=True, color="FFFFFF", size=11)
    even_fill      = PatternFill("solid", fgColor="EEF2F7")
    warn_fill      = PatternFill("solid", fgColor="FDE68A")  # ámbar — falta el dato
    no_match_fill  = PatternFill("solid", fgColor="FCA5A5")  # rojo — sin descripción, acción obligatoria
    invalid_fill   = PatternFill("solid", fgColor="DDD6FE")  # violeta — hay dato, pero no del tipo que esa columna espera

    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, r in enumerate(rows, 2):
        # Recalculado server-side a partir de los mismos campos _raw que
        # ya viajaron en el preview — no confía en el array "warnings" que
        # mandó el cliente, única fuente de verdad para el coloreado.
        warnings = _compute_warnings(r)
        zebra = even_fill if row_idx % 2 == 0 else None
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=r.get(field))
            cell.alignment = Alignment(vertical="center")
            warn_code = next((w for w, c in warn_col.items() if c == col_idx and w in warnings), None)
            if warn_code == "missing_description":
                cell.fill = no_match_fill
            elif warn_code in _INVALID_TYPE_CODES:
                cell.fill = invalid_fill
            elif warn_code:
                cell.fill = warn_fill
            elif zebra:
                cell.fill = zebra

    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
