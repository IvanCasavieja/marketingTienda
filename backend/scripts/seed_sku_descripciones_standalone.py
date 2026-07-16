"""
Script standalone de seed — usa asyncpg directo, sin dependencias de la app.
Carga Gestion/Diccionario.xlsx (SKU + "Descripción cenefa") en la tabla
sku_descripciones (migración 0028). Toca la base de PRODUCCIÓN directamente
-- por eso pide confirmación explícita salvo que se pase --yes.

Semántica de upsert: si el SKU no existe, INSERT. Si existe y nadie lo
corrigió nunca desde el Convertidor de Excel (updated_by_id IS NULL), se
UPDATEa con el valor del diccionario. Si existe y ya fue corregido por un
humano vía la herramienta (updated_by_id IS NOT NULL), se SALTEA -- no
queremos que una corrida futura de este script con un diccionario offline
desactualizado pise una corrección más reciente hecha en producción.

Uso:
  DATABASE_URL=postgresql+asyncpg://... python backend/scripts/seed_sku_descripciones_standalone.py [--dry-run] [--yes]
"""
import os
import sys
import re
import asyncio
import argparse

try:
    import openpyxl
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

try:
    import asyncpg
except ImportError:
    print("ERROR: pip install asyncpg")
    sys.exit(1)


def normalize_sku(raw) -> str | None:
    """int/float/str crudo de la celda SKU -> string canónico ('454520.0' -> '454520')."""
    if raw is None:
        return None
    if isinstance(raw, float):
        raw = int(raw) if raw.is_integer() else raw
    s = str(raw).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s or None


def parse_diccionario(path: str) -> list[tuple[str, str]]:
    """Lee columnas SKU (A) y Descripción cenefa (C). Descripción gestión
    (B) se ignora a propósito -- no se persiste, es la descripción mala."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(min_row=2, values_only=True)  # fila 1 = headers

    pares: list[tuple[str, str]] = []
    vistos: set[str] = set()
    for row in rows:
        if not row or len(row) < 3:
            continue
        sku = normalize_sku(row[0])
        descripcion = str(row[2]).strip() if row[2] is not None else ""
        if not sku or not descripcion:
            continue
        if sku in vistos:
            continue  # primer valor gana si el SKU está duplicado en el Excel
        vistos.add(sku)
        pares.append((sku, descripcion))
    return pares


async def run_seed(excel_path: str, dry_run: bool) -> None:
    raw_url = os.environ.get("DATABASE_URL", "")
    if not raw_url:
        print("ERROR: DATABASE_URL no definida")
        sys.exit(1)

    dsn = re.sub(r"^postgresql\+asyncpg://", "postgresql://", raw_url)
    dsn = re.sub(r"^postgres://", "postgresql://", dsn)

    pares = parse_diccionario(excel_path)
    print(f"Leídas {len(pares)} filas válidas (SKU + descripción no vacíos) de {excel_path}")

    conn = await asyncpg.connect(dsn)
    try:
        # Un solo round-trip para saber qué SKUs ya existen y si están
        # "bloqueados" (updated_by_id) -- antes esto era un SELECT por fila
        # (5000+ round-trips secuenciales contra Supabase, minutos de espera
        # para nada). executemany() más abajo también evita el patrón N+1
        # equivalente en la escritura.
        skus = [sku for sku, _ in pares]
        existing_rows = await conn.fetch(
            "SELECT sku, id, updated_by_id FROM sku_descripciones WHERE sku = ANY($1::text[])",
            skus,
        )
        existing = {r["sku"]: (r["id"], r["updated_by_id"]) for r in existing_rows}

        to_insert: list[tuple[str, str]] = []
        to_update: list[tuple[str, int]] = []  # (descripcion, id)
        skipped_locked = 0

        for sku, descripcion in pares:
            row = existing.get(sku)
            if row is None:
                to_insert.append((sku, descripcion))
            elif row[1] is None:  # updated_by_id
                to_update.append((descripcion, row[0]))
            else:
                skipped_locked += 1

        if not dry_run:
            if to_insert:
                await conn.executemany(
                    "INSERT INTO sku_descripciones (sku, descripcion) VALUES ($1, $2)",
                    to_insert,
                )
            if to_update:
                await conn.executemany(
                    "UPDATE sku_descripciones SET descripcion = $1, updated_at = now() WHERE id = $2",
                    to_update,
                )
    finally:
        await conn.close()

    prefix = "[DRY RUN] " if dry_run else ""
    print(
        f"\n{prefix}Listo — {len(to_insert)} insertados, {len(to_update)} actualizados, "
        f"{skipped_locked} salteados (ya corregidos a mano por un usuario)"
    )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "excel_path", nargs="?",
        default=os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "Gestion", "Diccionario.xlsx")),
        help="Ruta al Diccionario.xlsx (default: Gestion/Diccionario.xlsx en la raíz del repo)",
    )
    parser.add_argument("--dry-run", action="store_true", help="Solo simula, no escribe nada en la base")
    parser.add_argument("--yes", action="store_true", help="Saltea la confirmación interactiva")
    args = parser.parse_args()

    if not os.path.exists(args.excel_path):
        print(f"ERROR: no encontré el Excel en {args.excel_path}")
        sys.exit(1)

    if not args.dry_run and not args.yes:
        pares = parse_diccionario(args.excel_path)
        print(f"Esto va a escribir hasta {len(pares)} filas en sku_descripciones de la base de PRODUCCIÓN.")
        confirm = input("Escribí CONFIRMAR para continuar: ")
        if confirm.strip() != "CONFIRMAR":
            print("Cancelado.")
            sys.exit(0)

    print(f"Importando desde: {args.excel_path}")
    asyncio.run(run_seed(args.excel_path, args.dry_run))


if __name__ == "__main__":
    main()
