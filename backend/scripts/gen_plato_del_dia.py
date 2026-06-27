# -*- coding: utf-8 -*-
"""Genera Plato del dia - datos.xlsx con los datos de julio 2026."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DATA = [
    # (DIA,              DESCRIPCION,                                        PRECIO, SKU)
    ("MIERCOLES 1",  "TALLARINES C/TUCO DE CARNE UN.",                     179,  "36853"),
    ("JUEVES 2",     "PASTEL DE POLLO Y CALABAZA",                          279,  "37643"),
    ("VIERNES 3",    "MILANESA DE CARNE C/PURE",                            309,  "41122"),
    ("LUNES 6",      "STROGONOFF DE POLLO C/ARROZ UN.",                     189,  "72187"),
    ("MARTES 7",     "FELIPE JAMON Y QUESO",                                199,  "1332/14158"),
    ("MIERCOLES 8",  "RAVIOLES C/SALSA CARUSO UN.",                         199,  "12736"),
    ("JUEVES 9",     "MILANESA DE POLLO C/PURE UN.",                        249,  "52508"),
    ("VIERNES 10",   "FEIJOADA C/ARROZ UN.",                                215,  "17876"),
    ("LUNES 13",     "TARTA DE ZAPALLITOS INDIVIDUAL",                      239,  "57952"),
    ("MARTES 14",    "MUSLO DE POLLO C/RUSA",                               229,  "64610"),
    ("MIERCOLES 15", "SPAGHETTI C/ALBONDIGAS UN.",                          159,  "10383"),
    ("JUEVES 16",    "ENSALADA BALANCE UN.",                                 159,  "93678"),
    ("VIERNES 17",   "TORTA DE POLLO INDIVIDUAL",                           239,  "16819"),
    ("LUNES 20",     "GUISITO DE CARNE C/ARVEJAS Y ARROZ UN.",              299,  "65487"),
    ("MARTES 21",    "ALBONDIGAS C/ARROZ UN.",                              219,  "52155"),
    ("MIERCOLES 22", "RAVIOLES C/TUCO DE CARNE UN.",                        259,  "19272"),
    ("JUEVES 23",    "POLLO ARROLLADO C/ENSALADA RUSA UN.",                 285,  "80364"),
    ("VIERNES 24",   "CAZUELA DE LENTEJAS C/ARROZ UN.",                     199,  "50044"),
    ("LUNES 27",     "TARTA DE CEBOLLA INDIVIDUAL",                         215,  "57953"),
    ("MARTES 28",    "PECHUGA A LA PORTUGUESA C/ARROZ UN.",                 219,  "59915"),
    ("MIERCOLES 29", "ÑOQUIS C/TUCO DE CARNE",                         189,  "19870"),
    ("JUEVES 30",    "TORTA DE CARNE INDIVIDUAL",                           319,  "55462"),
    ("VIERNES 31",   "MERLUZA A LA MARINERA C/PURE",                        339,  "52170"),
]

HEADERS = ["DIA", "DESCRIPCION", "PRECIO", "ACLARACION", "DESCUENTO", "codigoSKU"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cenefas"

header_fill = PatternFill("solid", fgColor="1E3A5F")
header_font = Font(bold=True, color="FFFFFF", size=11)

for col, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

even_fill = PatternFill("solid", fgColor="EEF2F7")
for row_idx, (dia, desc, precio, sku) in enumerate(DATA, 2):
    fill = even_fill if row_idx % 2 == 0 else None
    for col_idx, val in enumerate([dia, desc, precio, None, False, sku], 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = Alignment(vertical="center")
        if fill:
            cell.fill = fill

ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 44
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 28
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 14
ws.row_dimensions[1].height = 26
ws.freeze_panes = "A2"

import pathlib
out = pathlib.Path(__file__).parent.parent.parent / "Plato del dia - datos.xlsx"
wb.save(out)
print(f"Guardado en: {out}")
print(f"Filas de datos: {len(DATA)}")
