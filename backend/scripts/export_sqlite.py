"""
export_sqlite.py — Exporta el SQLite del scraper a Excel.
Lee de la DB local (sin Postgres). Genera un Excel por día.

Genera en archivos/YYYY-MM-DD/:
  - precios_YYYY-MM-DD.xlsx   → una pestaña por cadena (GDU se divide en Disco/Devoto/Geant)
  - resumen_YYYY-MM-DD.txt    → totales por cadena

USO:
  python scripts/export_sqlite.py --fecha 2026-06-27 --db C:/tmp/scraper/productos.db
"""

import argparse
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

# Windows cp1252 fix — forzar UTF-8 en stdout
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_ROOT        = Path(__file__).resolve().parent.parent
ARCHIVOS_DIR = _ROOT.parent / "archivos"

COLUMNAS = [
    ("tienda",          "Cadena",          14),
    ("nombre",          "Nombre",          50),
    ("precio",          "Precio",          12),
    ("precio_lista",    "Precio Lista",    12),
    ("sku",             "SKU",             18),
    ("barcode",         "Barcode",         18),
    ("marca",           "Marca",           22),
    ("categoria",       "Categoría",       35),
    ("sucursal_id",     "ID Sucursal",     12),
    ("sucursal_nombre", "Sucursal",        30),
    ("url",             "URL",             55),
    ("actualizado_en",  "Actualizado",     20),
]

MAX_FILAS    = 1_000_000
COLOR_HEADER = "1F4E79"

ORDEN_CADENAS = ["Ta-Ta", "Farmashop", "Botiga", "ElDorado",
                 "Disco", "Devoto", "Geant"]


def _get_tiendas(con: sqlite3.Connection) -> list[str]:
    rows = con.execute("SELECT DISTINCT tienda FROM productos ORDER BY tienda").fetchall()
    return [r[0] for r in rows]


def _contar_por_tienda(con: sqlite3.Connection) -> dict[str, int]:
    rows = con.execute("SELECT tienda, COUNT(*) FROM productos GROUP BY tienda").fetchall()
    return dict(rows)


def _escribir_hoja(ws, keys: list[str], cursor: sqlite3.Cursor) -> int:
    """Escribe filas al worksheet en modo write_only. Retorna cantidad escrita."""
    n = 0
    for row in cursor:
        ws.append([row[i] for i in range(len(keys))])
        n += 1
        if n % 100_000 == 0:
            print(f"  ... {n:,} filas escritas", flush=True)
    return n


def _header(ws, labels: list[str], anchos: list[int]):
    fill = PatternFill("solid", fgColor=COLOR_HEADER)
    font = Font(bold=True, color="FFFFFF", size=11)
    header_row = []
    for i, (label, ancho) in enumerate(zip(labels, anchos), 1):
        cell = openpyxl.cell.WriteOnlyCell(ws, value=label)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        header_row.append(cell)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.append(header_row)
    ws.freeze_panes = "A2"


def exportar(fecha: str, db_path: Path):
    con = sqlite3.connect(db_path)
    conteos = _contar_por_tienda(con)
    if not conteos:
        print("SQLite vacío — nada que exportar.")
        con.close()
        return

    total = sum(conteos.values())
    print(f"Total: {total:,} filas en {len(conteos)} cadenas")

    dir_fecha = ARCHIVOS_DIR / fecha
    dir_fecha.mkdir(parents=True, exist_ok=True)

    keys   = [c[0] for c in COLUMNAS]
    labels = [c[1] for c in COLUMNAS]
    anchos = [c[2] for c in COLUMNAS]
    cols_sql = ", ".join(keys)

    xlsx_path = dir_fecha / f"precios_{fecha}.xlsx"
    print(f"\nGenerando {xlsx_path.name} ...")

    wb = openpyxl.Workbook(write_only=True)

    # Orden: cadenas conocidas primero, luego el resto
    tiendas_ord = [t for t in ORDEN_CADENAS if t in conteos]
    tiendas_ext = sorted(t for t in conteos if t not in ORDEN_CADENAS)

    for tienda in tiendas_ord + tiendas_ext:
        n_total = conteos[tienda]
        print(f"\n  {tienda}: {n_total:,} filas", flush=True)

        if n_total <= MAX_FILAS:
            ws = wb.create_sheet(title=tienda[:31])
            _header(ws, labels, anchos)
            cur = con.execute(
                f"SELECT {cols_sql} FROM productos WHERE tienda=? ORDER BY nombre",
                (tienda,)
            )
            escritas = _escribir_hoja(ws, keys, cur)
            print(f"    → pestaña '{tienda}': {escritas:,} filas")
        else:
            # Dividir en bloques de MAX_FILAS
            parte = 1
            offset = 0
            while offset < n_total:
                titulo = f"{tienda[:28]}_{parte}"
                ws = wb.create_sheet(title=titulo)
                _header(ws, labels, anchos)
                cur = con.execute(
                    f"SELECT {cols_sql} FROM productos WHERE tienda=? ORDER BY nombre "
                    f"LIMIT ? OFFSET ?",
                    (tienda, MAX_FILAS, offset)
                )
                escritas = _escribir_hoja(ws, keys, cur)
                print(f"    → pestaña '{titulo}': {escritas:,} filas")
                offset += MAX_FILAS
                parte  += 1

    wb.save(xlsx_path)
    con.close()

    kb = xlsx_path.stat().st_size // 1024
    print(f"\nExcel guardado: {xlsx_path} ({kb:,} KB)")

    # Resumen
    lines = [
        f"Resumen de precios — {fecha}",
        f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Total filas: {total:,}",
        "",
    ]
    for tienda in tiendas_ord + tiendas_ext:
        n = conteos[tienda]
        lines.append(f"  {tienda:<14}: {n:>9,} filas")
    resumen_path = dir_fecha / f"resumen_{fecha}.txt"
    resumen_path.write_text("\n".join(lines), encoding="utf-8")
    print("\n" + "\n".join(lines))


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--fecha", default=str(date.today()))
    ap.add_argument("--db",    default="C:/tmp/scraper/productos.db")
    args = ap.parse_args()

    db = Path(args.db)
    if not db.exists():
        print(f"ERROR: SQLite no encontrado: {db}")
        sys.exit(1)

    exportar(args.fecha, db)
