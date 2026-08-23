"""Deja solo el numero en las columnas de precio de una carpeta de planillas.

Los precios vienen con la palabra "unidad" pegada ("239 unidad") y en algunos
casos con una frase entera ("Comprando 2 $224,5 unidad"). El motor no puede
tratarlos como numero: los pasa tal cual al cartel, y el texto se desborda.

Por defecto simula. Con --aplicar escribe, dejando copia del original.
"""
import glob
import os
import pathlib
import re
import shutil
import sys

sys.path.insert(0, ".")

import openpyxl

from app.services.cenefas.variables import PRICE_VARS, resolve

D = pathlib.Path(sys.argv[1]) if len(sys.argv) > 1 and not sys.argv[1].startswith("--") else pathlib.Path.cwd()
RESPALDO = D / "originales"

# El precio es el numero que sigue al "$" cuando hay uno ("Comprando 2 $224,5
# unidad" -> 224,5); si no, el numero con el que arranca la celda ("239
# unidad" -> 239). Sin esa distincion, "Comprando 2 ..." devolveria 2.
_TRAS_PESOS = re.compile(r"\$\s*([\d.,]+)")
_AL_INICIO = re.compile(r"^\s*([\d.,]+)")


def limpiar(valor):
    if valor is None or isinstance(valor, (int, float)):
        return valor, False
    texto = str(valor).strip()
    if not texto:
        return valor, False
    m = _TRAS_PESOS.search(texto) or _AL_INICIO.match(texto)
    if not m:
        return valor, False
    limpio = m.group(1).strip().rstrip(".,")
    return (limpio, True) if limpio != texto else (valor, False)


aplicar = "--aplicar" in sys.argv
total = 0

for ruta in sorted(glob.glob(str(D / "Cenefas_MRP_*.xlsx"))):
    nombre = os.path.basename(ruta)
    wb = openpyxl.load_workbook(ruta)
    ws = wb.worksheets[0]

    filas = list(ws.iter_rows(values_only=True))
    hdr_i = next((i for i, r in enumerate(filas[:10])
                  if len([c for c in r if isinstance(c, str) and c.strip()]) >= 3), None)
    if hdr_i is None:
        print(f"{nombre}: no encontre la fila de encabezados"); continue

    cols = {c.column: str(c.value).strip()
            for c in ws[hdr_i + 1]
            if c.value is not None and resolve(str(c.value).strip()) in PRICE_VARS}
    if not cols:
        continue

    cambios = []
    for col, titulo in cols.items():
        for fila in range(hdr_i + 2, ws.max_row + 1):
            celda = ws.cell(row=fila, column=col)
            nuevo, cambio = limpiar(celda.value)
            if cambio:
                cambios.append((fila, titulo, str(celda.value), nuevo))
                if aplicar:
                    celda.value = nuevo

    if not cambios:
        print(f"{nombre}: ya esta limpia")
        continue

    print(f"{nombre}: {len(cambios)} celdas")
    for fila, titulo, antes, despues in cambios[:6]:
        print(f"    fila {fila:<4} {titulo:<16} {antes[:30]:<30} -> {despues}")
    if len(cambios) > 6:
        print(f"    ... y {len(cambios) - 6} mas")
    total += len(cambios)

    if aplicar:
        RESPALDO.mkdir(exist_ok=True)
        copia = RESPALDO / nombre
        if not copia.exists():          # nunca pisar el respaldo original
            shutil.copy2(ruta, copia)
        wb.save(ruta)

print()
print(f"{total} celdas de precio {'limpiadas' if aplicar else 'a limpiar'}.")
if not aplicar:
    print("SIMULACION -- correr con --aplicar para confirmar.")
