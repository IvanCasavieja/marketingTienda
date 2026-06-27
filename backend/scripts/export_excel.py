"""
export_excel.py — Exporta todos los productos de Supabase a un Excel.

Una hoja por cadena: Disco, Devoto, Geant, Ta-Ta, Farmashop, Botiga,
ElDorado (y cualquier otra tienda presente en la DB).

USO:
  python scripts/export_excel.py                      → precios_YYYY-MM-DD.xlsx
  python scripts/export_excel.py --out mi_archivo.xlsx
"""

import asyncio
import sys
import os
from datetime import date
from pathlib import Path

# ── bootstrap ─────────────────────────────────────────────────────────────────
_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT))
os.environ.setdefault("APP_ENV", "production")

from dotenv import load_dotenv
load_dotenv(_ROOT / ".env")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import sqlalchemy as sa
from sqlalchemy.ext.asyncio import create_async_engine


# ── Configuración ─────────────────────────────────────────────────────────────

COLUMNAS = [
    ("nombre",          "Nombre",           50),
    ("precio",          "Precio",           12),
    ("precio_lista",    "Precio Lista",     12),
    ("sku",             "SKU",              18),
    ("barcode",         "Barcode",          18),
    ("marca",           "Marca",            22),
    ("categoria",       "Categoría",        35),
    ("sucursal_id",     "ID Sucursal",      12),
    ("sucursal_nombre", "Sucursal",         30),
    ("url",             "URL",              55),
    ("actualizado_en",  "Actualizado",      20),
]

# Orden de hojas (si la cadena existe en la DB aparece en este orden)
ORDEN_CADENAS = ["Disco", "Devoto", "Geant", "Ta-Ta", "Farmashop", "Botiga", "ElDorado"]

COLOR_HEADER = "1F4E79"   # azul oscuro
COLOR_TEXT   = "FFFFFF"


# ── Helpers de estilo ─────────────────────────────────────────────────────────

def _estilo_header(ws, n_cols: int):
    fill = PatternFill("solid", fgColor=COLOR_HEADER)
    font = Font(bold=True, color=COLOR_TEXT, size=11)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20


def _ajustar_columnas(ws, anchos: list[int]):
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def _freeze_header(ws):
    ws.freeze_panes = "A2"


# ── Carga de datos ─────────────────────────────────────────────────────────────

async def _cargar_productos(engine) -> dict[str, list[dict]]:
    """Devuelve {tienda: [row_dict, ...]} ordenado por tienda."""
    query = sa.text("""
        SELECT
            tienda, nombre, precio, precio_lista, sku, barcode,
            marca, categoria, sucursal_id, sucursal_nombre, url,
            actualizado_en
        FROM productos
        ORDER BY tienda, nombre
    """)
    por_tienda: dict[str, list[dict]] = {}
    async with engine.connect() as conn:
        result = await conn.execute(query)
        for row in result.mappings():
            t = row["tienda"]
            por_tienda.setdefault(t, []).append(dict(row))
    return por_tienda


# ── Construcción del Excel ────────────────────────────────────────────────────

def _escribir_hoja(wb: openpyxl.Workbook, nombre_hoja: str, filas: list[dict]):
    ws = wb.create_sheet(title=nombre_hoja[:31])  # Excel limita a 31 chars

    # Cabecera
    keys    = [c[0] for c in COLUMNAS]
    labels  = [c[1] for c in COLUMNAS]
    anchos  = [c[2] for c in COLUMNAS]

    for col, label in enumerate(labels, 1):
        ws.cell(row=1, column=col, value=label)

    _estilo_header(ws, len(COLUMNAS))
    _ajustar_columnas(ws, anchos)
    _freeze_header(ws)

    # Datos
    for r, fila in enumerate(filas, 2):
        for col, key in enumerate(keys, 1):
            val = fila.get(key)
            # Formatear datetime
            if hasattr(val, "date"):
                val = val.strftime("%Y-%m-%d %H:%M")
            ws.cell(row=r, column=col, value=val)

    print(f"  {nombre_hoja}: {len(filas):,} productos")


# ── Main ──────────────────────────────────────────────────────────────────────

async def main(out_path: Path):
    db_url = os.environ.get("DATABASE_URL", "")
    if not db_url:
        print("ERROR: DATABASE_URL no configurada en .env")
        sys.exit(1)

    # asyncpg para leer, psycopg2-style para openpyxl (síncrono)
    engine = create_async_engine(db_url, pool_size=2, max_overflow=0)

    print("Conectando a Supabase y descargando productos...")
    por_tienda = await _cargar_productos(engine)
    await engine.dispose()

    total = sum(len(v) for v in por_tienda.values())
    print(f"Total productos en DB: {total:,} en {len(por_tienda)} cadenas\n")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja vacía por defecto

    # Escribir en orden definido primero, luego cualquier extra
    cadenas_ordenadas = [c for c in ORDEN_CADENAS if c in por_tienda]
    cadenas_extra     = sorted(t for t in por_tienda if t not in ORDEN_CADENAS)

    for cadena in cadenas_ordenadas + cadenas_extra:
        _escribir_hoja(wb, cadena, por_tienda[cadena])

    wb.save(out_path)
    print(f"\nExcel guardado en: {out_path}")
    print(f"Hojas: {', '.join(cadenas_ordenadas + cadenas_extra)}")


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=f"precios_{date.today()}.xlsx",
                    help="Archivo de salida (default: precios_YYYY-MM-DD.xlsx)")
    args = ap.parse_args()

    out = Path(args.out)
    asyncio.run(main(out))
