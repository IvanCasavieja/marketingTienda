"""
Script de importación de datos históricos del Excel Redexpress.
Uso: python backend/scripts/import_redexpress.py

Requiere: openpyxl, psycopg2 (o asyncpg via DATABASE_URL)
"""
import os
import sys
import re

# Add backend to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

try:
    import openpyxl
except ImportError:
    print("ERROR: instalar openpyxl: pip install openpyxl")
    sys.exit(1)

import asyncio
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker
from sqlalchemy import select
from app.models.planilla_pedido import PlanillaPedido

# ── Nombre canónico de locales ─────────────────────────────────────────────

CANONICAL_LOCALES = [
    "Nativo Florida",
    "Abast. Nativo B. Blancos (1000)",
    "Abast. Almenara (800)",
    "Abast. Ecomarket 3 (680)",
    "Abast. Del sol (780)",
    "Gatti (600)",
    "El Tio 1 (500)",
    "Frigo Yaro (400)",
    "Super 2 Hermanos (476)",
    "ALTOSUR 4 (545)",
    "SUPER 18 -1 (560)",
    "CARNETEL (500)",
    "Costa Verde (600)",
    "EXPRES 1 (480)",
    "El Morro (480)",
    "Fuentes (500)",
    "Nativo Suarez (400)",
    "MAROÑAS (366)",
    "Abast. San Ramón (393)",
    "Super Uno 1 (503)",
    "ABAST. SUPERMERCADO DONATO (380)",
    "Super Rodi (400)",
    "AVENIDA NORTE - SAN JOSE (600)",
    "ALTOSUR 1 (425)",
    "Frigo Centro (300)",
    "CARROUSELL (310)",
    "FLORESTA",
    "ARIEL 2 - MILLAN Y RAFFO (297)",
    "El Tio 2 (270)",
    "PINAMAR (290)",
    "Abast. La Cueva (300)",
    "El Tano (300)",
    "EXPRES 8 (330)",
    "Jardines (130)",
    "Abast. Hiperprecios (300)",
    "EXPRES 7 (295)",
    "Kampante (300)",
    "EXPRES 3 (250)",
    "EXPRES 6 (340)",
    "Comva (300)",
    "Abast. Santa Rosa (330)",
    "Super Uno 2 (270)",
    "AVENIDA MOLINO - SAN JOSE (200)",
    "AVENIDA SUR - SAN JOSE (300)",
    "SANTA CECILIA - SAN JOSE (225)",
    "SUPER 18 -2 (320)",
    "PAZ PLAZA - LA PAZ (300)",
    "L.A. DE HERRERA Y RAÑA (189)",
    "Prisma (228)",
    "DONATO EXPRESS (250)",
    "RED EXPRES NUEVO PARIS (245)",
    "AVENIDA CENTRO - SAN JOSE (200)",
    "ALTOSUR 6 (200)",
    "ALTOSUR 5 (230)",
    "OCHOA24 - SAN JOSE (100)",
    "ALTOSUR 2 (68)",
    "ALTOSUR 3 (90)",
    "PANDO",
    "JOY PANDO",
    "La Familia",
]

CANONICAL_SET = {loc.strip().lower(): loc for loc in CANONICAL_LOCALES}

# Variantes conocidas (de distintos meses del Excel)
ALIASES: dict[str, str] = {
    "nativo suarez(400)": "Nativo Suarez (400)",
    "super uno 1  (503)": "Super Uno 1 (503)",
    "floresta ": "FLORESTA",
    "la floresta - k53 (350)": "FLORESTA",
    "el tio2 (270)": "El Tio 2 (270)",
    "maroñas (366)": "MAROÑAS (366)",
    "abast. san ramón (393)": "Abast. San Ramón (393)",
    "l.a. de herrera y raña (189)": "L.A. DE HERRERA Y RAÑA (189)",
    "santa cecilia - playa pascual (200)": "SANTA CECILIA - SAN JOSE (225)",
    "avenida norte- san jose (600)": "AVENIDA NORTE - SAN JOSE (600)",
    "avenida molino- san jose (200)": "AVENIDA MOLINO - SAN JOSE (200)",
    "avenida sur- san jose (300)": "AVENIDA SUR - SAN JOSE (300)",
    "avenida centro- san jose (200)": "AVENIDA CENTRO - SAN JOSE (200)",
    "ochoa24- san jose (100)": "OCHOA24 - SAN JOSE (100)",
    "super 18 -1 (560)": "SUPER 18 -1 (560)",
    "expres 6": "EXPRES 6 (340)",
    "paz plaza - la paz (300)": "PAZ PLAZA - LA PAZ (300)",
}


def normalize_local(name: str | None) -> str | None:
    if not name or not isinstance(name, str):
        return None
    key = name.strip().lower()
    if key in ALIASES:
        return ALIASES[key]
    if key in CANONICAL_SET:
        return CANONICAL_SET[key]
    return None


# ── Mapeo de columnas Excel → campos DB ──────────────────────────────────────

# El mapeo depende del header de cada hoja
COLUMN_MAP = {
    "a4 oferta vertical":     "a4_oferta_vertical",
    "a4 oferta veritcal":     "a4_oferta_vertical",
    "cenefa oferta x3":       "cenefa_oferta_x3",
    "pinchos":                "pinchos",
    "pinchos x9":             "pinchos",
    "afiche 54x74":           "afiche_54x74",
    "cenefa valle del sol":   "cenefa_valle_del_sol",
    "cenefa supremo hogar":   "cenefa_supremo_hogar",
    "bombas 3x a4":           "bombas_3xa4",
    "bombas a4":              "bombas_a4",
    "bombas 74x 54":          "bombas_74x54",
    "bombas 74x54":           "bombas_74x54",
    "pinchos bombas":         "pinchos_bombas",
    "pinchos bombas x9":      "pinchos_bombas",
    "sticker valle del sol":  "sticker_valle_del_sol",
    "sticker carne":          "sticker_carne",
    "cenefas 3xa4 preciazos": "cenefas_preciazos",
    "afiche a4 super ahorro": "afiche_super_ahorro",
    "pinchos dias expres":    "pinchos_dias_expres",
    "hojas amarillas":        "hojas_amarillas",
    "otros":                  "otros",
    "otro":                   "otros",
}

MONTH_NAMES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}


def safe_int(val) -> int | None:
    if val is None:
        return None
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    if isinstance(val, str):
        val = val.strip()
        if val in ("", "NO", "no", "N/A"):
            return None
        try:
            return int(float(val))
        except ValueError:
            return None
    return None


def safe_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def parse_sheet(ws) -> tuple[int, int, list[dict]]:
    """Returns (year, month, list_of_row_dicts)."""
    rows = list(ws.iter_rows(values_only=True))

    # Detect month/year from first two rows
    year, month = 2026, 0
    for r in rows[:3]:
        for cell in r:
            if cell is None:
                continue
            if hasattr(cell, "year"):  # datetime
                year = cell.year
                month = cell.month
                break
            if isinstance(cell, str):
                low = cell.lower().strip()
                for mn, mv in MONTH_NAMES.items():
                    if mn in low:
                        month = mv
                        break
            if month:
                break
        if month:
            break

    if not month:
        return year, 0, []

    # Find header row (contains "Regionales" or "A4")
    header_row_idx = None
    col_map: dict[int, str] = {}
    for i, row in enumerate(rows):
        row_vals = [str(c).strip().lower() if c else "" for c in row]
        if any("a4 oferta" in v or "cenefa oferta" in v for v in row_vals):
            header_row_idx = i
            for j, val in enumerate(row_vals):
                if val in COLUMN_MAP:
                    col_map[j] = COLUMN_MAP[val]
            break

    if header_row_idx is None or not col_map:
        return year, month, []

    records = []
    for row in rows[header_row_idx + 1:]:
        if not row or row[0] is None:
            continue
        local_raw = str(row[0]).strip() if row[0] else None
        canonical = normalize_local(local_raw)
        if not canonical:
            continue  # skip totals row and unrecognized names

        data: dict = {"local_nombre": canonical}
        for col_idx, field in col_map.items():
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if field in ("otros", "hojas_amarillas"):
                data[field] = safe_str(val)
            else:
                data[field] = safe_int(val)

        records.append(data)

    return year, month, records


async def run_import(excel_path: str):
    raw_url = os.environ.get("DATABASE_URL", "")
    if not raw_url:
        print("ERROR: DATABASE_URL no definida")
        sys.exit(1)

    # Convert postgres:// → postgresql+asyncpg://
    db_url = re.sub(r"^postgres(ql)?://", "postgresql+asyncpg://", raw_url)

    engine = create_async_engine(db_url, echo=False)
    async_session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    total_upserted = 0

    async with async_session() as session:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            year, month, records = parse_sheet(ws)
            if not month:
                print(f"  ⚠  {sheet_name}: no se pudo detectar mes, saltando")
                continue

            print(f"  {sheet_name} → {year}/{month:02d} — {len(records)} locales")

            for rec in records:
                local = rec.pop("local_nombre")
                existing = await session.execute(
                    select(PlanillaPedido).where(
                        PlanillaPedido.local_nombre == local,
                        PlanillaPedido.year == year,
                        PlanillaPedido.month == month,
                    )
                )
                row = existing.scalar_one_or_none()

                if row is None:
                    row = PlanillaPedido(local_nombre=local, year=year, month=month)
                    session.add(row)

                for field, value in rec.items():
                    if value is not None:
                        setattr(row, field, value)

                total_upserted += 1

        await session.commit()

    print(f"\n✓ Importación completa — {total_upserted} filas insertadas/actualizadas")
    await engine.dispose()


if __name__ == "__main__":
    excel = os.path.join(
        os.path.dirname(__file__), "..", "..", "Pedido mensual 2026 cartelería.xlsx"
    )
    excel = os.path.abspath(excel)
    if not os.path.exists(excel):
        print(f"ERROR: no encontré el Excel en {excel}")
        sys.exit(1)

    print(f"Importando desde: {excel}")
    asyncio.run(run_import(excel))
