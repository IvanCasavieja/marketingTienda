"""
export_diario.py — Exporta el snapshot diario del scraper a archivos.

Genera en archivos/YYYY-MM-DD/:
  - precios_YYYY-MM-DD.xlsx        → Excel completo, una hoja por cadena
  - {Cadena}_YYYY-MM-DD.csv.gz     → CSV comprimido por cadena con todas las filas
  - resumen_YYYY-MM-DD.txt         → resumen de totales por cadena

Llamado automáticamente al final de cada scan completo.

USO MANUAL:
  python scripts/export_diario.py
  python scripts/export_diario.py --fecha 2026-06-27
  python scripts/export_diario.py --solo-csv
  python scripts/export_diario.py --solo-excel
"""

import argparse
import asyncio
import csv
import gzip
import io
import os
import sys
from datetime import date, datetime
from pathlib import Path

# ── bootstrap ────────────────────────────────────────────────────────────────
_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT))
os.environ.setdefault("APP_ENV", "production")
from dotenv import load_dotenv
load_dotenv(_ROOT / ".env")

import sqlalchemy as sa
from sqlalchemy.ext.asyncio import create_async_engine
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Configuración ─────────────────────────────────────────────────────────────

ARCHIVOS_DIR = _ROOT.parent / "archivos"

COLUMNAS = [
    ("nombre",          "Nombre",          50),
    ("precio",          "Precio",          12),
    ("precio_lista",    "Precio Lista",    12),
    ("sku",             "SKU",             18),
    ("barcode",         "Barcode",         18),
    ("marca",           "Marca",           22),
    ("categoria",       "Categoría",       35),
    ("sucursal_id",     "ID Sucursal",     12),
    ("sucursal_nombre", "Sucursal",        30),
    ("url",             "URL",             55),
    ("actualizado_en",  "Actualizado",     20),
]

ORDEN_CADENAS = ["Disco", "Devoto", "Geant", "Ta-Ta", "ElDorado",
                 "Farmashop", "Botiga", "Pigalle"]

COLOR_HEADER = "1F4E79"

# ── Carga de datos ────────────────────────────────────────────────────────────

async def _cargar(engine) -> dict[str, list[dict]]:
    query = sa.text("""
        SELECT tienda, nombre, precio, precio_lista, sku, barcode,
               marca, categoria, sucursal_id, sucursal_nombre, url, actualizado_en
        FROM productos
        ORDER BY tienda, nombre
    """)
    por_tienda: dict[str, list[dict]] = {}
    async with engine.connect() as conn:
        for row in (await conn.execute(query)).mappings():
            por_tienda.setdefault(row["tienda"], []).append(dict(row))
    return por_tienda

# ── CSV.gz por cadena ─────────────────────────────────────────────────────────

def _escribir_csv_gz(path: Path, filas: list[dict]):
    keys = [c[0] for c in COLUMNAS]
    headers = [c[1] for c in COLUMNAS]
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for fila in filas:
        w.writerow([
            fila.get(k).strftime("%Y-%m-%d %H:%M") if hasattr(fila.get(k), "strftime") else fila.get(k)
            for k in keys
        ])
    path.write_bytes(gzip.compress(buf.getvalue().encode("utf-8"), compresslevel=9))

# ── Excel ─────────────────────────────────────────────────────────────────────

def _hoja(wb: openpyxl.Workbook, nombre: str, filas: list[dict]):
    """Escribe una hoja. Si supera 1M filas, divide en _1, _2, etc."""
    MAX_FILAS = 1_000_000

    def _crear_hoja(titulo: str, chunk: list[dict]):
        ws = wb.create_sheet(title=titulo[:31])
        keys   = [c[0] for c in COLUMNAS]
        labels = [c[1] for c in COLUMNAS]
        anchos = [c[2] for c in COLUMNAS]

        fill = PatternFill("solid", fgColor=COLOR_HEADER)
        font = Font(bold=True, color="FFFFFF", size=11)
        for col, label in enumerate(labels, 1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 20
        for i, w in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        for r, fila in enumerate(chunk, 2):
            for col, key in enumerate(keys, 1):
                val = fila.get(key)
                if hasattr(val, "strftime"):
                    val = val.strftime("%Y-%m-%d %H:%M")
                ws.cell(row=r, column=col, value=val)

    if len(filas) <= MAX_FILAS:
        _crear_hoja(nombre, filas)
    else:
        for i, start in enumerate(range(0, len(filas), MAX_FILAS), 1):
            _crear_hoja(f"{nombre[:28]}_{i}", filas[start:start + MAX_FILAS])

# ── Resumen ───────────────────────────────────────────────────────────────────

def _resumen(dir_fecha: Path, por_tienda: dict, fecha: str):
    total = sum(len(v) for v in por_tienda.values())
    lines = [
        f"Resumen de precios — {fecha}",
        f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Total filas: {total:,}",
        "",
    ]
    for cadena in ORDEN_CADENAS + sorted(t for t in por_tienda if t not in ORDEN_CADENAS):
        if cadena in por_tienda:
            n = len(por_tienda[cadena])
            skus = len({r.get("sku") for r in por_tienda[cadena] if r.get("sku")})
            suc  = len({r.get("sucursal_id") for r in por_tienda[cadena] if r.get("sucursal_id")})
            lines.append(f"  {cadena:<12}: {n:>8,} filas | {skus:>7,} SKUs únicos | {suc:>3} sucursales")
    (dir_fecha / f"resumen_{fecha}.txt").write_text("\n".join(lines), encoding="utf-8")
    print("\n".join(lines))

# ── Main ──────────────────────────────────────────────────────────────────────

async def main(fecha: str, solo_csv: bool, solo_excel: bool):
    db_url = os.environ.get("DATABASE_URL", "")
    if not db_url:
        print("ERROR: DATABASE_URL no configurada")
        sys.exit(1)

    engine = create_async_engine(db_url, pool_size=2, max_overflow=0)
    print(f"Descargando datos de Supabase...")
    por_tienda = await _cargar(engine)
    await engine.dispose()

    total = sum(len(v) for v in por_tienda.values())
    print(f"Total: {total:,} filas en {len(por_tienda)} cadenas")

    dir_fecha = ARCHIVOS_DIR / fecha
    dir_fecha.mkdir(parents=True, exist_ok=True)

    cadenas_ord = [c for c in ORDEN_CADENAS if c in por_tienda]
    cadenas_ext = sorted(t for t in por_tienda if t not in ORDEN_CADENAS)

    # ── CSV.gz por cadena ──────────────────────────────────────────────────
    if not solo_excel:
        print("\nGenerando CSV.gz...")
        for cadena in cadenas_ord + cadenas_ext:
            filas = por_tienda[cadena]
            path  = dir_fecha / f"{cadena}_{fecha}.csv.gz"
            _escribir_csv_gz(path, filas)
            kb = path.stat().st_size // 1024
            print(f"  {cadena}: {len(filas):,} filas → {kb:,} KB comprimido")

    # ── Excel ──────────────────────────────────────────────────────────────
    if not solo_csv:
        print("\nGenerando Excel...")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for cadena in cadenas_ord + cadenas_ext:
            filas = por_tienda[cadena]
            _hoja(wb, cadena, filas)
            print(f"  {cadena}: {len(filas):,} filas")
        xlsx_path = dir_fecha / f"precios_{fecha}.xlsx"
        wb.save(xlsx_path)
        kb = xlsx_path.stat().st_size // 1024
        print(f"Excel: {xlsx_path.name} ({kb:,} KB)")

    # ── Resumen ────────────────────────────────────────────────────────────
    _resumen(dir_fecha, por_tienda, fecha)
    print(f"\nArchivos en: {dir_fecha}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--fecha",      default=str(date.today()), help="YYYY-MM-DD")
    ap.add_argument("--solo-csv",   action="store_true")
    ap.add_argument("--solo-excel", action="store_true")
    args = ap.parse_args()
    asyncio.run(main(args.fecha, args.solo_csv, args.solo_excel))
